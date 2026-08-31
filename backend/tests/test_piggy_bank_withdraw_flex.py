"""
Unit tests for POST /api/owner/piggy-bank/withdraw (flexible «Снять на расходы»)
and the dedicated piggy-bank Excel export.

Covers:
- Withdrawal from a chosen piggy bank (wash/detailing) without bookingId
- source="piggy": withdrawal + Expense mirror, NO payroll entries (no ЗП link)
- source="own": own-money expense + salary compensation (PayrollEntry bonus),
  no piggy transaction, piggy balance untouched
- expenseCategory: custom budget category honored
- withdrawKind="other" → other_withdrawal transaction + Expense «Прочие расходы» (legacy)
- Validation: no target → 400, invalid resourceGroup → 422, worker → 403
- Legacy contract: withdrawal with bookingId still resolves the bucket from the booking
- Balances: balance AND combinedBalance both drop by the withdrawn amount
- GET /api/owner/exports/piggy-bank builds an xlsx with Сводка/Операции sheets
- comparison sheets «По месяцам»/«По неделям» (сб–пт) with period bounds, deltas, balances
- GET /api/owner/exports/report contains a «Копилка» sheet with our transaction
"""
from __future__ import annotations

import io
import json
import os
import sys
import unittest
import urllib.parse
from pathlib import Path
from uuid import uuid4

from fastapi.testclient import TestClient


def reset_app_modules() -> None:
    for name in list(sys.modules):
        if (
            name == "app"
            or name.startswith(("app.", "backend.app."))
            or name == "backend.app"
            or name == "bot"
        ):
            del sys.modules[name]


def build_init_data(telegram_id: str) -> str:
    """Build Telegram init data that passes insecure validation (no HMAC)."""
    return urllib.parse.urlencode({"user": json.dumps({"id": int(telegram_id)})})


class PiggyBankWithdrawFlexTests(unittest.TestCase):
    OWNER_TG_ID = "777951"
    WORKER_TG_ID = "777953"

    def setUp(self) -> None:
        data_dir = Path(__file__).resolve().parents[1] / "data"
        data_dir.mkdir(parents=True, exist_ok=True)
        self.db_path = data_dir / f"test_suite_{uuid4().hex}.sqlite3"
        os.environ["DATABASE_URL"] = f"sqlite:///{self.db_path.as_posix()}"
        os.environ["APP_ENV"] = "development"
        os.environ["APP_SECRET"] = "test-secret"
        os.environ["CRON_SECRET"] = "test-cron-secret"
        os.environ["ALLOW_DEMO_SEED_DATA"] = "true"
        os.environ["RUN_EMBEDDED_BOT"] = "false"
        os.environ["ALLOW_INSECURE_CLIENT_AUTH"] = "true"
        os.environ["TELEGRAM_BOT_TOKEN"] = "123456:test-bot-token"
        os.environ["TELEGRAM_DELIVERY_MODE"] = "polling"
        os.environ["SYNC_TELEGRAM_WEBHOOK"] = "false"
        os.environ["TELEGRAM_WEBHOOK_PATH"] = "/api/telegram/webhook"
        os.environ.pop("WEBAPP_URL", None)

        reset_app_modules()
        from app.main import app

        self.client_manager = TestClient(app)
        self.client = self.client_manager.__enter__()

        self._set_staff_telegram_ids()
        self.owner_token = build_init_data(self.OWNER_TG_ID)
        self.worker_token = build_init_data(self.WORKER_TG_ID)

    def tearDown(self) -> None:
        if hasattr(self, "client_manager"):
            self.client_manager.__exit__(None, None, None)
        try:
            from app.database import engine
        except ModuleNotFoundError:
            pass
        else:
            engine.dispose()
        reset_app_modules()
        if self.db_path.exists():
            self.db_path.unlink()

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _set_staff_telegram_ids(self) -> None:
        from app.database import SessionLocal
        from app.models import StaffUser
        from sqlalchemy import select

        with SessionLocal() as db:
            owner = db.scalar(select(StaffUser).where(StaffUser.login == "owner"))
            worker = db.scalar(select(StaffUser).where(StaffUser.role == "worker"))
            if owner is not None:
                owner.telegram_chat_id = self.OWNER_TG_ID
            if worker is not None:
                worker.telegram_chat_id = self.WORKER_TG_ID
            db.commit()

    @staticmethod
    def _auth_headers(token: str) -> dict[str, str]:
        return {"Authorization": token}

    def _withdraw(self, **payload) -> tuple[int, dict]:
        response = self.client.post(
            "/api/owner/piggy-bank/withdraw",
            headers=self._auth_headers(self.owner_token),
            json=payload,
        )
        return response.status_code, (
            response.json() if response.content else {}
        )

    def _piggy_bank(self) -> dict:
        response = self.client.get(
            "/api/owner/piggy-bank",
            headers=self._auth_headers(self.owner_token),
        )
        self.assertEqual(response.status_code, 200, response.text)
        return response.json()

    def _create_booking_with_service(self, resource_group: str = "wash") -> str:
        """Insert Client + Service + Booking directly; return booking id."""
        from app.database import SessionLocal
        from app.models import Booking, Client, Service

        suffix = uuid4().hex[:10]
        with SessionLocal() as db:
            client = Client(
                id=f"c-{suffix}",
                name="Клиент Копилка",
                phone=f"+7999000{suffix[:4]}",
                car="BMW",
                plate=f"A{suffix[:3]}BC",
            )
            service = Service(
                id=f"s-{suffix}",
                name="Мойка тест",
                category="Мойка",
                price=2000,
                duration=60,
                resource_group=resource_group,
            )
            db.add_all([client, service])
            db.flush()
            booking = Booking(
                id=f"b-{suffix}",
                client_id=client.id,
                client_name=client.name,
                client_phone=client.phone,
                service=service.name,
                service_id=service.id,
                date="01.08.2026",
                time="10:00",
                duration=service.duration,
                price=service.price,
                status="completed",
                box="Бокс 1",
                payment_type="cash",
            )
            db.add(booking)
            db.commit()
            return booking.id

    # ------------------------------------------------------------------
    # Tests: withdrawal without booking binding
    # ------------------------------------------------------------------

    def test_withdraw_without_booking_from_detailing(self) -> None:
        before = self._piggy_bank()
        status_code, tx = self._withdraw(
            resourceGroup="detailing",
            materialName="Пленка PPF",
            materialCost=3000,
            purpose="На детейлинг-заказ",
            date="10.08.2026",
        )
        self.assertEqual(status_code, 200)
        self.assertEqual(tx["transactionType"], "material_withdrawal")
        self.assertIsNone(tx["bookingId"])
        self.assertEqual(tx["resourceGroup"], "detailing")
        self.assertEqual(tx["amount"], -3000)
        self.assertEqual(tx["purpose"], "На детейлинг-заказ")

        after = self._piggy_bank()
        self.assertEqual(after["balance"], before["balance"] - 3000)
        self.assertEqual(after["combinedBalance"], before["combinedBalance"] - 3000)

    def test_withdraw_other_kind_creates_other_expense(self) -> None:
        before = self._piggy_bank()
        status_code, tx = self._withdraw(
            resourceGroup="wash",
            materialName="Ремонт пылесоса",
            materialCost=2500,
            purpose="Прочие расходы на мойку",
            date="11.08.2026",
            withdrawKind="other",
        )
        self.assertEqual(status_code, 200)
        self.assertEqual(tx["transactionType"], "other_withdrawal")
        self.assertEqual(tx["resourceGroup"], "wash")
        self.assertEqual(tx["source"], "piggy")

        after = self._piggy_bank()
        self.assertEqual(after["balance"], before["balance"] - 2500)
        self.assertEqual(
            after["remainingInPiggyBank"],
            before["remainingInPiggyBank"] - 2500,
        )

        from app.database import SessionLocal
        from app.models import Expense
        from sqlalchemy import select

        with SessionLocal() as db:
            expense = db.scalar(
                select(Expense).where(Expense.title == "Прочие расходы: Ремонт пылесоса")
            )
            self.assertIsNotNone(expense)
            assert expense is not None
            self.assertEqual(expense.category, "Прочие расходы")
            self.assertEqual(float(expense.amount), 2500)
    def test_piggy_source_creates_no_payroll_entries(self) -> None:
        """Снятие из копилки не должно трогать зарплату (без удержаний)."""
        from app.database import SessionLocal
        from app.models import PayrollEntry
        from sqlalchemy import func, select

        with SessionLocal() as db:
            payroll_before = db.scalar(select(func.count(PayrollEntry.id)))

        status_code, tx = self._withdraw(
            resourceGroup="detailing",
            materialName="Полироль",
            materialCost=1500,
            date="18.08.2026",
            source="piggy",
            spentByName="Иван Тест",
        )
        self.assertEqual(status_code, 200)
        self.assertIsNone(tx["payrollEntryId"])

        with SessionLocal() as db:
            payroll_after = db.scalar(select(func.count(PayrollEntry.id)))
            self.assertEqual(payroll_after, payroll_before, "снятие из копилки создало запись в ЗП")

    def test_own_money_expense_compensates_in_salary(self) -> None:
        """Свои деньги: расход бюджета + компенсация (bonus) в ЗП, копилка не трогается.

        Без «кто взял» компенсация уходит текущему пользователю (fallback на actor).
        """
        from app.database import SessionLocal
        from app.models import Expense, PayrollEntry, PiggyBankTransaction, StaffUser
        from sqlalchemy import func, select

        with SessionLocal() as db:
            owner = db.scalar(select(StaffUser).where(StaffUser.telegram_chat_id == self.OWNER_TG_ID))
            assert owner is not None
            owner_id = owner.id

        before = self._piggy_bank()

        status_code, tx = self._withdraw(
            resourceGroup="wash",
            materialName="Химия за свой счёт",
            materialCost=2000,
            purpose="Купил сам, компенсируйте",
            date="19.08.2026",
            source="own",
        )
        self.assertEqual(status_code, 200, tx)
        self.assertEqual(tx["transactionType"], "own_expense")
        self.assertEqual(tx["source"], "own")
        self.assertIsNone(tx["bookingId"])
        self.assertIsNotNone(tx["payrollEntryId"], "компенсация не создана")

        # Копилка не изменилась (транзакций нет; combinedBalance может меняться —
        # существующая формула вычитает ЛЮБЫЕ расходы сегмента из остатка)
        after = self._piggy_bank()
        self.assertEqual(after["balance"], before["balance"])
        self.assertEqual(after["detailing"]["netPiggy"], before["detailing"]["netPiggy"])

        with SessionLocal() as db:
            # Нет зеркальной транзакции копилки
            mirror_count = db.scalar(
                select(func.count(PiggyBankTransaction.id)).where(
                    PiggyBankTransaction.purpose == "Купил сам, компенсируйте"
                )
            )
            self.assertEqual(mirror_count, 0, "own-расход попал в историю копилки")

            # Расход бюджета
            expense = db.scalar(
                select(Expense).where(Expense.title == "Материалы: Химия за свой счёт")
            )
            self.assertIsNotNone(expense)
            assert expense is not None
            self.assertEqual(float(expense.amount), 2000)

            # Компенсация в ЗП (bonus, привязан к расходу, у текущего пользователя)
            comp = db.get(PayrollEntry, tx["payrollEntryId"])
            self.assertIsNotNone(comp, "компенсация в ЗП не создана")
            assert comp is not None
            self.assertEqual(comp.kind, "bonus")
            self.assertEqual(float(comp.amount), 2000)
            self.assertEqual(comp.worker_id, owner_id)
            self.assertEqual(comp.entry_date, "19.08.2026")
            self.assertEqual(comp.expense_id, expense.id)

    def test_own_money_without_person_returns_consistent_response(self) -> None:
        """Свои деньги: ответ консистентен (компенсация уходит actor-fallback'у)."""
        status_code, tx = self._withdraw(
            resourceGroup="detailing",
            materialName="Что-то за наличку",
            materialCost=800,
            date="20.08.2026",
            source="own",
        )
        self.assertEqual(status_code, 200)
        self.assertEqual(tx["source"], "own")

    def test_custom_expense_category_honored(self) -> None:
        from app.database import SessionLocal
        from app.models import Expense
        from sqlalchemy import select

        status_code, _ = self._withdraw(
            resourceGroup="wash",
            materialName="Реклама",
            materialCost=3000,
            date="21.08.2026",
            source="own",
            expenseCategory="Маркетинг",
        )
        self.assertEqual(status_code, 200)

        with SessionLocal() as db:
            expense = db.scalar(
                select(Expense).where(Expense.title == "Материалы: Реклама")
            )
            self.assertIsNotNone(expense)
            assert expense is not None
            self.assertEqual(expense.category, "Маркетинг")

    def test_own_money_expense_edit_updates_compensation(self) -> None:
        """Правка расхода обновляет и компенсацию (PATCH /api/expenses/{id})."""
        from app.database import SessionLocal
        from app.models import PayrollEntry

        status_code, tx = self._withdraw(
            resourceGroup="wash",
            materialName="Лампы",
            materialCost=1000,
            date="22.08.2026",
            source="own",
        )
        self.assertEqual(status_code, 200)
        assert tx["payrollEntryId"] is not None

        expense_id = None
        with SessionLocal() as db:
            comp = db.get(PayrollEntry, tx["payrollEntryId"])
            assert comp is not None
            expense_id = comp.expense_id
            self.assertIsNotNone(expense_id)

        # Редактируем расход бюджета на 1300
        patch_response = self.client.patch(
            f"/api/expenses/{expense_id}",
            headers=self._auth_headers(self.owner_token),
            json={"amount": 1300},
        )
        self.assertEqual(patch_response.status_code, 200, patch_response.text)

        with SessionLocal() as db:
            comp = db.get(PayrollEntry, tx["payrollEntryId"])
            assert comp is not None
            self.assertEqual(float(comp.amount), 1300, "компенсация не синхронизировалась с расходом")

    def test_withdraw_other_default_purpose(self) -> None:
        status_code, tx = self._withdraw(
            resourceGroup="detailing",
            materialName="Химия для химчистки",
            materialCost=1200,
            date="12.08.2026",
            withdrawKind="other",
        )
        self.assertEqual(status_code, 200)
        self.assertEqual(tx["purpose"], "Прочие расходы: Химия для химчистки")

    def test_without_target_rejected(self) -> None:
        status_code, body = self._withdraw(
            materialName="Что-то",
            materialCost=500,
            date="13.08.2026",
        )
        self.assertEqual(status_code, 400)
        self.assertIn("копилку", body.get("detail", ""))

    def test_invalid_resource_group_rejected(self) -> None:
        status_code, _ = self._withdraw(
            resourceGroup="carwash",
            materialName="Что-то",
            materialCost=500,
            date="13.08.2026",
        )
        self.assertEqual(status_code, 422)

    def test_worker_forbidden(self) -> None:
        response = self.client.post(
            "/api/owner/piggy-bank/withdraw",
            headers=self._auth_headers(self.worker_token),
            json={
                "resourceGroup": "wash",
                "materialName": "Химия",
                "materialCost": 500,
                "date": "13.08.2026",
            },
        )
        self.assertEqual(response.status_code, 403, response.text)

    # ------------------------------------------------------------------
    # Tests: legacy contract with bookingId keeps working
    # ------------------------------------------------------------------

    def test_legacy_withdraw_with_booking_resolves_bucket(self) -> None:
        booking_id = self._create_booking_with_service(resource_group="wash")
        status_code, tx = self._withdraw(
            bookingId=booking_id,
            materialName="Воск",
            materialCost=700,
            purpose="Для записи",
            date="14.08.2026",
        )
        self.assertEqual(status_code, 200)
        self.assertEqual(tx["bookingId"], booking_id)
        self.assertEqual(tx["resourceGroup"], "wash")
        self.assertEqual(tx["transactionType"], "material_withdrawal")

    # ------------------------------------------------------------------
    # Tests: Excel exports
    # ------------------------------------------------------------------

    def test_piggy_export_endpoint_returns_xlsx(self) -> None:
        self._withdraw(
            resourceGroup="detailing",
            materialName="Полироль",
            materialCost=900,
            purpose="Экспорт-тест",
            date="15.08.2026",
            withdrawKind="other",
        )
        response = self.client.get(
            "/api/owner/exports/piggy-bank",
            headers=self._auth_headers(self.owner_token),
            params={"date_from": "01.08.2026", "date_to": "31.08.2026"},
        )
        self.assertEqual(response.status_code, 200, response.text)
        self.assertTrue(response.headers["Content-Disposition"].startswith("attachment"))
        self.assertIn("piggy-bank-report", response.headers["Content-Disposition"])

        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(response.content))
        self.assertIn("Сводка", workbook.sheetnames)
        self.assertIn("Операции", workbook.sheetnames)
        operations = workbook["Операции"]
        flat_rows = [
            [cell.value for cell in row] for row in operations.iter_rows()
        ]
        purposes = [str(row[3]) for row in flat_rows if row and row[3]]
        self.assertIn("Экспорт-тест", "".join(purposes))

    def test_piggy_export_comparison_sheets(self) -> None:
        # 15.08.2026 и 05.09.2026 — субботы: начало финансовой недели (сб–пт)
        self._withdraw(
            resourceGroup="wash",
            materialName="Химия",
            materialCost=1000,
            purpose="Сравнение·август",
            date="15.08.2026",
        )
        self._withdraw(
            resourceGroup="detailing",
            materialName="Паста",
            materialCost=2500,
            purpose="Сравнение·сентябрь",
            date="05.09.2026",
        )
        response = self.client.get(
            "/api/owner/exports/piggy-bank",
            headers=self._auth_headers(self.owner_token),
            params={"date_from": "01.08.2026", "date_to": "30.09.2026"},
        )
        self.assertEqual(response.status_code, 200, response.text)

        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(response.content))
        self.assertIn("По месяцам", workbook.sheetnames)
        self.assertIn("По неделям", workbook.sheetnames)

        # ── Сводка: новые строки ──
        summary_values = [
            str(cell.value)
            for row in workbook["Сводка"].iter_rows()
            for cell in row
            if cell.value is not None
        ]
        joined_summary = " | ".join(summary_values)
        self.assertIn("Операций за период", joined_summary)
        self.assertIn("Среднемесячное сальдо", joined_summary)
        self.assertIn("Лучший месяц по сальдо", joined_summary)

        # ── По месяцам ──
        month_rows = [
            [cell.value for cell in row]
            for row in workbook["По месяцам"].iter_rows()
        ]
        self.assertEqual(month_rows[0][1], "Начало месяца")
        self.assertEqual(month_rows[0][2], "Конец месяца")
        labels = [str(r[0]) for r in month_rows[1:] if r and r[0]]
        self.assertIn("Август 2026", labels)
        self.assertIn("Сентябрь 2026", labels)
        self.assertIn("ИТОГО", labels)

        august = next(r for r in month_rows[1:] if r[0] == "Август 2026")
        self.assertEqual(august[1], "01.08.2026")
        self.assertEqual(august[2], "31.08.2026")
        self.assertEqual(august[3], 1)                    # операций
        self.assertEqual(float(august[9]), 1000.0)        # снятие на материалы
        self.assertEqual(float(august[13]), -1000.0)      # сальдо периода
        self.assertIn(august[14], ("", None))             # Δ у первого месяца нет
        self.assertEqual(august[15], "—")
        self.assertEqual(float(august[16]), -1000.0)      # баланс на 31.08
        self.assertEqual(float(august[17]), -1000.0)      # сальдо · мойка
        self.assertEqual(float(august[18]), 0.0)          # сальдо · детейлинг

        september = next(r for r in month_rows[1:] if r[0] == "Сентябрь 2026")
        self.assertEqual(float(september[13]), -2500.0)
        self.assertEqual(float(september[14]), -1500.0)   # Δ к августу: -2500 − (−1000)
        self.assertEqual(september[15], "-150,0%")        # Δ% к прошлому месяцу
        self.assertEqual(float(september[16]), -3500.0)   # баланс на 30.09
        self.assertEqual(float(september[18]), -2500.0)   # сальдо · детейлинг

        itog = month_rows[-1]
        self.assertEqual(itog[0], "ИТОГО")
        self.assertEqual(itog[3], 2)
        self.assertEqual(float(itog[13]), -3500.0)
        self.assertEqual(float(itog[16]), -3500.0)

        # ── По неделям: начало/конец недели (сб–пт) ──
        week_rows = [
            [cell.value for cell in row]
            for row in workbook["По неделям"].iter_rows()
        ]
        self.assertEqual(week_rows[0][1], "Начало недели (сб)")
        self.assertEqual(week_rows[0][2], "Конец недели (пт)")

        week_aug = next(r for r in week_rows[1:] if r[1] == "15.08.2026")
        self.assertEqual(week_aug[2], "21.08.2026")
        self.assertEqual(float(week_aug[13]), -1000.0)
        self.assertEqual(float(week_aug[16]), -1000.0)

        week_sep = next(r for r in week_rows[1:] if r[1] == "05.09.2026")
        self.assertEqual(week_sep[2], "11.09.2026")
        self.assertEqual(float(week_sep[13]), -2500.0)
        self.assertEqual(float(week_sep[16]), -3500.0)

        # Пустые недели внутри периода тоже присутствуют (непрерывная сетка сравнения)
        week_empty = next(r for r in week_rows[1:] if r[1] == "01.08.2026")
        self.assertEqual(week_empty[3], 0)
        self.assertEqual(float(week_empty[13]), 0.0)

    def test_piggy_export_resource_group_filter(self) -> None:
        # Только копилка мойки: детейлинг-операции не попадают в отчёт
        self._withdraw(
            resourceGroup="wash",
            materialName="Химия",
            materialCost=1000,
            purpose="Только мойка",
            date="15.08.2026",
        )
        self._withdraw(
            resourceGroup="detailing",
            materialName="Паста",
            materialCost=2500,
            purpose="Только детейлинг",
            date="16.08.2026",
        )
        response = self.client.get(
            "/api/owner/exports/piggy-bank",
            headers=self._auth_headers(self.owner_token),
            params={"resource_group": "wash"},
        )
        self.assertEqual(response.status_code, 200, response.text)
        self.assertIn("piggy-bank-wash", response.headers["Content-Disposition"])

        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(response.content))
        operations = [
            [cell.value for cell in row]
            for row in workbook["Операции"].iter_rows()
        ]
        purposes = "".join(str(r[3]) for r in operations[1:] if r and r[3])
        self.assertIn("Только мойка", purposes)
        self.assertNotIn("Только детейлинг", purposes)

        summary_values = " | ".join(
            str(cell.value)
            for row in workbook["Сводка"].iter_rows()
            for cell in row
            if cell.value is not None
        )
        self.assertIn("Отчёт по копилке · Мойка", summary_values)

        bad = self.client.get(
            "/api/owner/exports/piggy-bank",
            headers=self._auth_headers(self.owner_token),
            params={"resource_group": "carwash"},
        )
        self.assertEqual(bad.status_code, 422, bad.text)

    def test_owner_report_contains_piggy_sheet(self) -> None:
        self._withdraw(
            resourceGroup="detailing",
            materialName="Аппликатор",
            materialCost=400,
            purpose="Лист копилки",
            date="16.08.2026",
        )
        response = self.client.get(
            "/api/owner/exports/report",
            headers=self._auth_headers(self.owner_token),
        )
        self.assertEqual(response.status_code, 200, response.text)

        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(response.content))
        self.assertIn("Копилка", workbook.sheetnames)
        sheet = workbook["Копилка"]
        values = [str(cell.value) for row in sheet.iter_rows() for cell in row]
        joined = " | ".join(values)
        self.assertIn("Снятие на материалы", joined)
        self.assertIn("Лист копилки", joined)

    def test_piggy_export_forbidden_for_worker(self) -> None:
        response = self.client.get(
            "/api/owner/exports/piggy-bank",
            headers=self._auth_headers(self.worker_token),
        )
        self.assertEqual(response.status_code, 403, response.text)


if __name__ == "__main__":
    unittest.main()
