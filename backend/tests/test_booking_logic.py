from __future__ import annotations

import os
import sys
import unittest
import json
import hmac
import hashlib
import time
from io import BytesIO
from datetime import datetime, timedelta, timezone
from pathlib import Path
from unittest.mock import patch
from uuid import uuid4

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select


def reset_app_modules() -> None:
    for name in list(sys.modules):
        if name == "app" or name.startswith("app."):
            del sys.modules[name]


class BookingLogicTests(unittest.TestCase):
    def setUp(self) -> None:
        data_dir = Path(__file__).resolve().parents[1] / "data"
        data_dir.mkdir(parents=True, exist_ok=True)
        self.db_path = data_dir / f"test_suite_{uuid4().hex}.sqlite3"
        os.environ["DATABASE_URL"] = f"sqlite:///{self.db_path.as_posix()}"
        os.environ["APP_SECRET"] = "test-secret"
        os.environ["RUN_EMBEDDED_BOT"] = "false"
        os.environ["ALLOW_INSECURE_CLIENT_AUTH"] = "true"
        os.environ["TELEGRAM_BOT_TOKEN"] = "123456:test-bot-token"
        os.environ["TELEGRAM_DELIVERY_MODE"] = "polling"
        os.environ["SYNC_TELEGRAM_WEBHOOK"] = "false"
        os.environ["TELEGRAM_WEBHOOK_PATH"] = "/api/telegram/webhook"
        os.environ.pop("WEBAPP_URL", None)

        self.restart_app()

    def tearDown(self) -> None:
        self.shutdown_app()
        reset_app_modules()
        if self.db_path.exists():
            self.db_path.unlink()

    def shutdown_app(self) -> None:
        if hasattr(self, "client_manager"):
            self.client_manager.__exit__(None, None, None)
        try:
            from app.database import engine
        except ModuleNotFoundError:
            return
        engine.dispose()

    def restart_app(self) -> None:
        if hasattr(self, "client_manager"):
            self.shutdown_app()
        reset_app_modules()
        from app.main import app

        self.client_manager = TestClient(app)
        self.client = self.client_manager.__enter__()

    def login_client(self, *, name: str, phone: str, car: str = "Lada Vesta", plate: str = "A123BC") -> tuple[str, str]:
        response = self.client.post("/api/auth/client", json=self.client_auth_payload(name=name, phone=phone, car=car, plate=plate))
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        return payload["token"], payload["bootstrap"]["session"]["actorId"]

    def client_auth_payload(
        self,
        *,
        name: str,
        phone: str,
        car: str = "Lada Vesta",
        plate: str = "A123BC",
        telegram_id: str | None = None,
    ) -> dict[str, object]:
        payload: dict[str, object] = {
            "profile": {
                "name": name,
                "phone": phone,
                "car": car,
                "plate": plate,
                "registered": True,
            }
        }
        if telegram_id is not None:
            payload["initData"] = self.make_init_data(telegram_id)
        return payload

    def make_init_data(
        self,
        telegram_id: str,
        *,
        first_name: str = "Alice",
        username: str | None = None,
        auth_date: int | None = None,
    ) -> str:
        user = {"id": int(telegram_id), "first_name": first_name}
        if username:
            user["username"] = username
        pairs = {
            "auth_date": str(int(time.time()) if auth_date is None else auth_date),
            "query_id": f"AAH{telegram_id}",
            "user": json.dumps(user, separators=(",", ":"), ensure_ascii=False),
        }
        data_check_string = "\n".join(f"{key}={pairs[key]}" for key in sorted(pairs))
        secret_key = hmac.new(b"WebAppData", os.environ["TELEGRAM_BOT_TOKEN"].encode("utf-8"), hashlib.sha256).digest()
        pairs["hash"] = hmac.new(secret_key, data_check_string.encode("utf-8"), hashlib.sha256).hexdigest()
        return "&".join(f"{key}={value}" for key, value in pairs.items())

    def telegram_webhook_secret(self) -> str:
        raw = f"{os.environ['APP_SECRET']}:{os.environ['TELEGRAM_BOT_TOKEN']}".encode("utf-8")
        return hashlib.sha256(raw).hexdigest()

    def login_staff(self, login: str, password: str) -> str:
        response = self.client.post(
            "/api/auth/staff/login",
            json={"login": login, "password": password},
        )
        self.assertEqual(response.status_code, 200, response.text)
        return response.json()["token"]

    def get_staff(self, *, login: str | None = None, staff_id: str | None = None) -> dict[str, object]:
        from app.database import SessionLocal
        from app.models import StaffUser

        if login is None and staff_id is None:
            raise AssertionError("Either login or staff_id must be provided")
        with SessionLocal() as db:
            staff = db.get(StaffUser, staff_id) if staff_id is not None else db.scalar(select(StaffUser).where(StaffUser.login == login))
            self.assertIsNotNone(staff)
            assert staff is not None
            return {
                "id": staff.id,
                "login": staff.login,
                "role": staff.role,
                "active": staff.active,
                "telegram_chat_id": staff.telegram_chat_id,
            }

    def get_client(self, client_id: str) -> dict[str, object]:
        from app.database import SessionLocal
        from app.models import Client

        with SessionLocal() as db:
            client = db.get(Client, client_id)
            self.assertIsNotNone(client)
            assert client is not None
            return {
                "id": client.id,
                "name": client.name,
                "phone": client.phone,
                "telegram_id": client.telegram_id,
                "car": client.car,
                "plate": client.plate,
            }

    def count_clients(self) -> int:
        from app.database import SessionLocal
        from app.models import Client

        with SessionLocal() as db:
            return len(db.scalars(select(Client)).all())

    def count_client_notifications(self, client_id: str) -> int:
        from app.database import SessionLocal
        from app.models import Notification

        with SessionLocal() as db:
            return len(
                db.scalars(
                    select(Notification).where(
                        Notification.recipient_role == "client",
                        Notification.recipient_id == client_id,
                    )
                ).all()
            )

    def count_client_sessions(self, client_id: str) -> int:
        from app.database import SessionLocal
        from app.models import AuthSession

        with SessionLocal() as db:
            return len(
                db.scalars(
                    select(AuthSession).where(
                        AuthSession.actor_role == "client",
                        AuthSession.actor_id == client_id,
                    )
                ).all()
            )

    def test_session_schema_supports_prefixed_ids_and_long_mobile_user_agents(self) -> None:
        from app.models import AuthSession, Booking, BookingWorker, Client, Expense, Notification, Penalty, StaffUser, StockItem, TelegramLinkCode

        self.assertEqual(getattr(Client.__table__.c.id.type, "length", None), 64)
        self.assertEqual(getattr(StaffUser.__table__.c.id.type, "length", None), 64)
        self.assertEqual(getattr(Booking.__table__.c.client_id.type, "length", None), 64)
        self.assertEqual(getattr(Booking.__table__.c.id.type, "length", None), 64)
        self.assertEqual(getattr(BookingWorker.__table__.c.booking_id.type, "length", None), 64)
        self.assertEqual(getattr(BookingWorker.__table__.c.worker_id.type, "length", None), 64)
        self.assertEqual(getattr(AuthSession.__table__.c.actor_id.type, "length", None), 64)
        self.assertIsNone(getattr(AuthSession.__table__.c.user_agent.type, "length", None))
        self.assertEqual(getattr(Notification.__table__.c.id.type, "length", None), 64)
        self.assertEqual(getattr(Notification.__table__.c.recipient_id.type, "length", None), 64)
        self.assertEqual(getattr(StockItem.__table__.c.id.type, "length", None), 64)
        self.assertEqual(getattr(Expense.__table__.c.id.type, "length", None), 64)
        self.assertEqual(getattr(Penalty.__table__.c.id.type, "length", None), 64)
        self.assertEqual(getattr(Penalty.__table__.c.worker_id.type, "length", None), 64)
        self.assertEqual(getattr(Penalty.__table__.c.owner_id.type, "length", None), 64)
        self.assertEqual(getattr(Penalty.__table__.c.revoked_by.type, "length", None), 64)
        self.assertEqual(getattr(TelegramLinkCode.__table__.c.staff_id.type, "length", None), 64)

        response = self.client.post(
            "/api/auth/client",
            headers={
                "user-agent": (
                    "Mozilla/5.0 (Linux; Android 13; K) AppleWebKit/537.36 "
                    "(KHTML, like Gecko) Chrome/146.0.7680.164 Mobile Safari/537.36 "
                    "Telegram-Android/12.1.1 (Realme RMX3363; Android 13; SDK 33; HIGH)"
                )
            },
            json=self.client_auth_payload(name="Alice", phone="+7 (999) 111-22-33"),
        )
        self.assertEqual(response.status_code, 200, response.text)
        actor_id = response.json()["actorId"]
        self.assertTrue(actor_id.startswith("c-"))
        self.assertGreaterEqual(self.count_client_sessions(actor_id), 1)

    def count_worker_notifications(self, worker_id: str) -> int:
        from app.database import SessionLocal
        from app.models import Notification

        with SessionLocal() as db:
            return len(
                db.scalars(
                    select(Notification).where(
                        Notification.recipient_role == "worker",
                        Notification.recipient_id == worker_id,
                    )
                ).all()
            )

    def disable_owner_two_factor(self) -> None:
        from app.database import SessionLocal
        from app.models import AppSetting

        with SessionLocal() as db:
            setting = db.get(AppSetting, "owner_security")
            self.assertIsNotNone(setting)
            assert setting is not None
            setting.value = {"twoFactor": False}
            db.commit()

    def test_secondary_owner_can_login_without_primary_owner_telegram_when_2fa_cannot_run(self) -> None:
        response = self.client.post(
            "/api/auth/staff/login",
            json={"login": "owner", "password": "owner"},
        )
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["role"], "owner")
        self.assertEqual(payload["bootstrap"]["session"]["role"], "owner")

    def set_primary_owner_telegram(self, chat_id: str = "974738256") -> None:
        from app.database import SessionLocal
        from app.models import StaffUser

        with SessionLocal() as db:
            owner = db.scalar(select(StaffUser).where(StaffUser.is_primary_owner.is_(True)))
            self.assertIsNotNone(owner)
            assert owner is not None
            owner.telegram_chat_id = chat_id
            db.commit()

    def set_staff_telegram(self, login: str, chat_id: str) -> None:
        from app.database import SessionLocal
        from app.models import StaffUser

        with SessionLocal() as db:
            staff = db.scalar(select(StaffUser).where(StaffUser.login == login))
            self.assertIsNotNone(staff)
            assert staff is not None
            staff.telegram_chat_id = chat_id
            db.commit()

    @staticmethod
    def extract_owner_reset_code(message: str) -> str:
        prefixes = ["Код подтверждения: ", "РљРѕРґ РїРѕРґС‚РІРµСЂР¶РґРµРЅРёСЏ: "]
        for line in message.splitlines():
            for prefix in prefixes:
                if line.startswith(prefix):
                    return line[len(prefix) :].strip()
        raise AssertionError(f"Owner reset code not found in message: {message}")

    def force_owner_reset_ready(self) -> None:
        from app.database import SessionLocal
        from app.models import AppSetting

        with SessionLocal() as db:
            setting = db.get(AppSetting, "owner_database_reset")
            self.assertIsNotNone(setting)
            assert setting is not None
            next_value = dict(setting.value or {})
            next_value["finalizeAfter"] = (datetime.now().astimezone() - timedelta(seconds=1)).isoformat()
            setting.value = next_value
            db.commit()

    @staticmethod
    def auth_headers(token: str) -> dict[str, str]:
        return {"Authorization": f"Bearer {token}"}

    @staticmethod
    def next_active_date() -> str:
        candidate = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        for offset in range(1, 8):
            next_date = candidate + timedelta(days=offset)
            if next_date.weekday() != 6:
                return next_date.strftime("%d.%m.%Y")
        raise AssertionError("Unable to find active schedule day")

    def test_client_booking_uses_session_client_and_forces_new_status(self) -> None:
        token, actor_id = self.login_client(name="Alice", phone="+7 (999) 111-22-33")
        response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(token),
            json={
                "clientId": "spoofed-client",
                "clientName": "Mallory",
                "clientPhone": "+7 (999) 111-22-33",
                "service": "Spoofed service",
                "serviceId": "s1",
                "date": self.next_active_date(),
                "time": "10:00",
                "duration": 999,
                "price": 1,
                "status": "completed",
                "workers": [{"workerId": "w1", "workerName": "????", "percent": 30}],
                "box": "???? 1",
                "paymentType": "cash",
                "car": "BMW",
                "plate": "M001AA",
            },
        )
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["clientId"], actor_id)
        self.assertEqual(payload["clientName"], "Alice")
        self.assertEqual(payload["status"], "new")
        self.assertEqual(payload["workers"], [])
        self.assertEqual(payload["serviceId"], "s1")
        self.assertNotEqual(payload["service"], "Spoofed service")
        self.assertNotEqual(payload["duration"], 999)

    def test_owner_can_update_client_card_notes_and_debt(self) -> None:
        _client_token, client_id = self.login_client(name="Alice", phone="+7 (999) 222-33-44")
        self.disable_owner_two_factor()
        owner_token = self.login_staff("owner", "owner")

        response = self.client.patch(
            f"/api/clients/{client_id}/card",
            headers=self.auth_headers(owner_token),
            json={"notes": "VIP РєР»РёРµРЅС‚", "debtBalance": 1500},
        )
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["notes"], "VIP РєР»РёРµРЅС‚")
        self.assertEqual(payload["debtBalance"], 1500)

        from app.database import SessionLocal
        from app.models import Client

        with SessionLocal() as db:
            client = db.get(Client, client_id)
            self.assertIsNotNone(client)
            assert client is not None
            self.assertEqual(client.notes, "VIP РєР»РёРµРЅС‚")
            self.assertEqual(client.debt_balance, 1500)

    def test_owner_dispatches_booking_reminders_once_per_booking(self) -> None:
        auth_response = self.client.post(
            "/api/auth/client",
            json=self.client_auth_payload(
                name="Alice",
                phone="+7 (999) 555-44-33",
                telegram_id="555111222",
            ),
        )
        self.assertEqual(auth_response.status_code, 200, auth_response.text)
        client_id = auth_response.json()["bootstrap"]["session"]["actorId"]

        self.disable_owner_two_factor()
        self.set_staff_telegram("ivan", "200200200")
        owner_token = self.login_staff("owner", "owner")
        reminder_date = self.next_active_date()

        from app.database import SessionLocal
        from app.models import AppSetting, Notification

        with SessionLocal() as db:
            worker_settings = db.get(AppSetting, "worker_notification_settings")
            self.assertIsNotNone(worker_settings)
            assert worker_settings is not None
            next_value = dict(worker_settings.value or {})
            next_value["w1"] = {
                "newTask": True,
                "taskUpdate": True,
                "payment": True,
                "reminders": True,
                "sms": False,
            }
            worker_settings.value = next_value
            db.commit()

        booking_response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(owner_token),
            json={
                "clientId": client_id,
                "clientName": "Alice",
                "clientPhone": "+7 (999) 555-44-33",
                "service": "РњРѕР№РєР° Р±Р°Р·РѕРІР°СЏ",
                "serviceId": "s1",
                "date": reminder_date,
                "time": "11:00",
                "duration": 30,
                "price": 1200,
                "status": "confirmed",
                "workers": [{"workerId": "w1", "workerName": "РРІР°РЅ", "percent": 30}],
                "box": "Р‘РѕРєСЃ 1",
                "paymentType": "cash",
                "car": "Lada Vesta",
                "plate": "A123BC",
            },
        )
        self.assertEqual(booking_response.status_code, 200, booking_response.text)

        sent_messages: list[tuple[str, str]] = []

        def fake_send_message(chat_id: str, text: str) -> None:
            sent_messages.append((chat_id, text))

        with patch("app.main.send_telegram_message", side_effect=fake_send_message):
            dispatch_response = self.client.post(
                "/api/owner/reminders/dispatch",
                headers=self.auth_headers(owner_token),
                json={"targetDate": reminder_date},
            )
            self.assertEqual(dispatch_response.status_code, 200, dispatch_response.text)
            dispatch_payload = dispatch_response.json()
            self.assertEqual(dispatch_payload["clientReminders"], 1)
            self.assertEqual(dispatch_payload["workerReminders"], 1)
            self.assertEqual(dispatch_payload["telegramDelivered"], 2)

            second_response = self.client.post(
                "/api/owner/reminders/dispatch",
                headers=self.auth_headers(owner_token),
                json={"targetDate": reminder_date},
            )
            self.assertEqual(second_response.status_code, 200, second_response.text)
            second_payload = second_response.json()
            self.assertEqual(second_payload["clientReminders"], 0)
            self.assertEqual(second_payload["workerReminders"], 0)
            self.assertEqual(second_payload["telegramDelivered"], 0)

        with SessionLocal() as db:
            client_notifications = db.scalars(
                select(Notification).where(
                    Notification.recipient_role == "client",
                    Notification.recipient_id == client_id,
                    Notification.message.like("%РќР°РїРѕРјРёРЅР°РЅРёРµ Рѕ Р·Р°РїРёСЃРё%"),
                )
            ).all()
            worker_notifications = db.scalars(
                select(Notification).where(
                    Notification.recipient_role == "worker",
                    Notification.recipient_id == "w1",
                    Notification.message.like("%РќР°РїРѕРјРёРЅР°РЅРёРµ РјР°СЃС‚РµСЂСѓ%"),
                )
            ).all()
            self.assertEqual(len(client_notifications), 1)
            self.assertEqual(len(worker_notifications), 1)

    def test_client_login_tolerates_legacy_partial_settings(self) -> None:
        from app.database import SessionLocal
        from app.models import AppSetting

        with SessionLocal() as db:
            owner_notifications = db.get(AppSetting, "owner_notification_settings")
            self.assertIsNotNone(owner_notifications)
            assert owner_notifications is not None
            owner_notifications.value = {
                "telegramBot": True,
                "emailReports": True,
                "smsReminders": False,
                "lowStock": True,
                "dailyReport": True,
                "weeklyReport": False,
            }

            worker_notifications = db.get(AppSetting, "worker_notification_settings")
            self.assertIsNotNone(worker_notifications)
            assert worker_notifications is not None
            worker_notifications.value = {
                "w1": {
                    "newTask": True,
                    "taskUpdate": True,
                    "payment": True,
                }
            }
            db.commit()

        response = self.client.post(
            "/api/auth/client",
            json=self.client_auth_payload(name="Alice", phone="+7 (999) 222-33-44"),
        )
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["bootstrap"]["settings"]["ownerNotificationSettings"]["bookingReminders"], False)
        self.assertEqual(payload["bootstrap"]["settings"]["workerNotificationSettings"], {})

    def test_client_booking_uses_other_active_box_when_first_is_busy(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        booking_date = self.next_active_date()
        admin_response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(admin_token),
            json={
                "clientId": "",
                "clientName": "First Client",
                "clientPhone": "+7 (999) 111-22-33",
                "service": "РњРѕР№РєР° Р±Р°Р·РѕРІР°СЏ",
                "serviceId": "s1",
                "date": booking_date,
                "time": "10:00",
                "duration": 30,
                "price": 1200,
                "status": "scheduled",
                "workers": [],
                "box": "Р‘РѕРєСЃ 1",
                "paymentType": "cash",
                "car": "Lada Vesta",
                "plate": "A123BC",
            },
        )
        self.assertEqual(admin_response.status_code, 200, admin_response.text)

        client_token, _actor_id = self.login_client(name="Alice", phone="+7 (999) 222-33-44", car="BMW X5", plate="A123BC")
        client_response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(client_token),
            json={
                "clientId": "",
                "clientName": "Ignored Name",
                "clientPhone": "+7 (999) 222-33-44",
                "service": "РњРѕР№РєР° Р±Р°Р·РѕРІР°СЏ",
                "serviceId": "s1",
                "date": booking_date,
                "time": "10:00",
                "duration": 30,
                "price": 1200,
                "status": "scheduled",
                "workers": [],
                "box": "Р‘РѕРєСЃ 1",
                "paymentType": "cash",
                "car": "BMW X5",
                "plate": "A123BC",
            },
        )
        self.assertEqual(client_response.status_code, 200, client_response.text)
        payload = client_response.json()
        self.assertNotEqual(payload["box"], "Р‘РѕРєСЃ 1")

    def test_detailing_booking_uses_detailing_room_and_keeps_slots_separate(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        booking_date = self.next_active_date()

        wash_response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(admin_token),
            json={
                "clientId": "",
                "clientName": "Wash Client",
                "clientPhone": "+7 (999) 111-22-33",
                "service": "Мойка базовая",
                "serviceId": "s1",
                "date": booking_date,
                "time": "11:00",
                "duration": 30,
                "price": 1200,
                "status": "scheduled",
                "workers": [],
                "box": "Бокс 1",
                "paymentType": "cash",
                "car": "Lada Vesta",
                "plate": "A123BC",
            },
        )
        self.assertEqual(wash_response.status_code, 200, wash_response.text)

        client_token, _ = self.login_client(name="Alice", phone="+7 (999) 222-33-44", car="BMW X5", plate="A123BC")
        availability_response = self.client.get(
            f"/api/bookings/availability?date={booking_date}&duration=60&serviceId=s2",
            headers=self.auth_headers(client_token),
        )
        self.assertEqual(availability_response.status_code, 200, availability_response.text)
        availability_payload = availability_response.json()
        slot = next(item for item in availability_payload["slots"] if item["time"] == "11:00")
        self.assertEqual(slot["freeBoxes"], 1)
        self.assertTrue(slot["available"])

        detailing_response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(client_token),
            json={
                "clientId": "",
                "clientName": "Ignored Name",
                "clientPhone": "+7 (999) 222-33-44",
                "service": "Полировка стекла",
                "serviceId": "s2",
                "date": booking_date,
                "time": "11:00",
                "duration": 60,
                "price": 3500,
                "status": "scheduled",
                "workers": [],
                "box": "Бокс 1",
                "paymentType": "cash",
                "car": "BMW X5",
                "plate": "A123BC",
            },
        )
        self.assertEqual(detailing_response.status_code, 200, detailing_response.text)
        detailing_payload = detailing_response.json()
        self.assertEqual(detailing_payload["box"], "Детейлинг")
        self.assertEqual(detailing_payload["status"], "new")

    def test_booking_rejects_box_time_overlap(self) -> None:
        token, _ = self.login_client(name="Alice", phone="+7 (999) 111-22-33")
        common = {
            "clientId": "",
            "clientName": "Alice",
            "clientPhone": "+7 (999) 111-22-33",
            "service": "????? + ?????????",
            "serviceId": "s5",
            "date": self.next_active_date(),
            "duration": 90,
            "price": 4200,
            "status": "scheduled",
            "workers": [],
            "box": "???? 1",
            "paymentType": "cash",
            "car": "Lada Vesta",
            "plate": "A123BC",
        }
        first = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(token),
            json={**common, "time": "10:00"},
        )
        self.assertEqual(first.status_code, 200, first.text)

        second = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(token),
            json={**common, "time": "10:30"},
        )
        self.assertEqual(second.status_code, 409, second.text)

    def test_booking_must_fit_schedule_window(self) -> None:
        token, _ = self.login_client(name="Alice", phone="+7 (999) 111-22-33")
        response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(token),
            json={
                "clientId": "",
                "clientName": "Alice",
                "clientPhone": "+7 (999) 111-22-33",
                "service": "????? + ?????????",
                "serviceId": "s5",
                "date": self.next_active_date(),
                "time": "20:30",
                "duration": 90,
                "price": 4200,
                "status": "scheduled",
                "workers": [],
                "box": "???? 1",
                "paymentType": "cash",
                "car": "Lada Vesta",
                "plate": "A123BC",
            },
        )
        self.assertEqual(response.status_code, 400, response.text)

    def test_worker_cannot_update_foreign_booking(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        create_response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(admin_token),
            json={
                "clientId": "",
                "clientName": "Alice",
                "clientPhone": "+7 (999) 111-22-33",
                "service": "????? ???????",
                "serviceId": "s1",
                "date": self.next_active_date(),
                "time": "11:00",
                "duration": 30,
                "price": 1200,
                "status": "scheduled",
                "workers": [],
                "box": "???? 1",
                "paymentType": "cash",
                "car": "Lada Vesta",
                "plate": "A123BC",
            },
        )
        self.assertEqual(create_response.status_code, 200, create_response.text)
        booking_id = create_response.json()["id"]

        worker_token = self.login_staff("ivan", "master")
        update_response = self.client.patch(
            f"/api/bookings/{booking_id}",
            headers=self.auth_headers(worker_token),
            json={"status": "in_progress"},
        )
        self.assertEqual(update_response.status_code, 403, update_response.text)

    def test_owner_can_revoke_all_worker_complaints(self) -> None:
        self.disable_owner_two_factor()
        owner_token = self.login_staff("owner", "owner")
        worker = self.get_staff(login="ivan")

        for title in ("РћРїРѕР·РґР°РЅРёРµ", "РљР°С‡РµСЃС‚РІРѕ"):
            response = self.client.post(
                "/api/penalties",
                headers=self.auth_headers(owner_token),
                json={
                    "workerId": worker["id"],
                    "title": title,
                    "reason": "РџСЂРѕРІРµСЂРєР° СЃРЅСЏС‚РёСЏ РІСЃРµС… Р¶Р°Р»РѕР±",
                },
            )
            self.assertEqual(response.status_code, 200, response.text)

        revoke_response = self.client.post(
            f"/api/workers/{worker['id']}/penalties/revoke-all",
            headers=self.auth_headers(owner_token),
        )
        self.assertEqual(revoke_response.status_code, 200, revoke_response.text)

        bootstrap_response = self.client.get("/api/auth/session", headers=self.auth_headers(owner_token))
        self.assertEqual(bootstrap_response.status_code, 200, bootstrap_response.text)
        worker_penalties = [item for item in bootstrap_response.json()["penalties"] if item["workerId"] == worker["id"]]
        self.assertTrue(worker_penalties)
        self.assertEqual(len([item for item in worker_penalties if item["revokedAt"] is None]), 0)

    def test_owner_summary_report_sends_detailed_excel_document(self) -> None:
        from app.database import SessionLocal
        from app.models import Booking, BookingWorker

        self.disable_owner_two_factor()
        self.set_primary_owner_telegram("123456789")
        owner_token = self.login_staff("owner", "owner")
        _client_token, client_id = self.login_client(name="Alice", phone="+7 (999) 111-22-33")
        today = datetime.now().strftime("%d.%m.%Y")
        created_at = datetime.now()

        with SessionLocal() as db:
            wash_booking = Booking(
                id=f"b-{uuid4()}",
                client_id=client_id,
                client_name="Alice",
                client_phone="+7 (999) 111-22-33",
                service="РњРѕР№РєР° Р±Р°Р·РѕРІР°СЏ",
                service_id="s1",
                date=today,
                time="10:00",
                duration=30,
                price=1200,
                status="completed",
                box="Р‘РѕРєСЃ 1",
                payment_type="cash",
                notes="РџРµРЅР° Рё СЃСѓС€РєР°",
                car="Lada Vesta",
                plate="A123BC",
                created_at=created_at,
            )
            detail_booking = Booking(
                id=f"b-{uuid4()}",
                client_id=client_id,
                client_name="Alice",
                client_phone="+7 (999) 111-22-33",
                service="РџРѕР»РёСЂРѕРІРєР° СЃС‚РµРєР»Р°",
                service_id="s2",
                date=today,
                time="11:00",
                duration=60,
                price=3500,
                status="completed",
                box="Р‘РѕРєСЃ 2",
                payment_type="card",
                notes="РќРµ РґРѕР»Р¶РЅРѕ РїРѕРїР°СЃС‚СЊ РІ РјРѕР№РєСѓ",
                car="Lada Vesta",
                plate="A123BC",
                created_at=created_at,
            )
            db.add_all([wash_booking, detail_booking])
            db.flush()
            db.add(
                BookingWorker(
                    booking_id=wash_booking.id,
                    worker_id="w2",
                    worker_name="РћР»РµРі",
                    percent=10,
                )
            )
            db.commit()

        sent_documents: list[dict[str, object]] = []

        def fake_send_document(chat_id: str | int, *, file_name: str, content: bytes, caption: str | None = None, mime_type: str = "application/octet-stream") -> None:
            sent_documents.append(
                {
                    "chat_id": str(chat_id),
                    "file_name": file_name,
                    "content": content,
                    "caption": caption or "",
                    "mime_type": mime_type,
                }
            )

        with patch("app.main.send_telegram_document", side_effect=fake_send_document):
            response = self.client.post(
                "/api/owner/reports/daily/wash/telegram",
                headers=self.auth_headers(owner_token),
            )

        self.assertEqual(response.status_code, 200, response.text)
        self.assertIn(".xlsx", response.json()["message"])
        self.assertEqual(len(sent_documents), 1)
        document = sent_documents[0]
        self.assertEqual(document["chat_id"], "123456789")
        self.assertTrue(str(document["file_name"]).endswith(".xlsx"))
        self.assertEqual(document["mime_type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIn("Р•Р¶РµРґРЅРµРІРЅС‹Р№ РѕС‚С‡С‘С‚ РїРѕ РЅР°РїСЂР°РІР»РµРЅРёСЋ: РњРѕР№РєР°", str(document["caption"]))

        workbook = load_workbook(filename=BytesIO(document["content"]))
        self.assertIn("РЎРІРѕРґРєР°", workbook.sheetnames)
        self.assertIn("РЎС‚Р°С‚СѓСЃС‹", workbook.sheetnames)
        self.assertIn("РЎРѕС‚СЂСѓРґРЅРёРєРё", workbook.sheetnames)
        self.assertIn("Р РµРµСЃС‚СЂ Р·Р°РїРёСЃРµР№", workbook.sheetnames)

        summary = workbook["РЎРІРѕРґРєР°"]
        self.assertEqual(summary["A2"].value, "Р•Р¶РµРґРЅРµРІРЅС‹Р№ РѕС‚С‡С‘С‚ РїРѕ РЅР°РїСЂР°РІР»РµРЅРёСЋ: РњРѕР№РєР°")

        registry_rows = list(workbook["Р РµРµСЃС‚СЂ Р·Р°РїРёСЃРµР№"].iter_rows(min_row=2, values_only=True))
        meaningful_registry = [row for row in registry_rows if row[0] != "РќРµС‚ РґР°РЅРЅС‹С…"]
        self.assertEqual(len(meaningful_registry), 1)
        self.assertEqual(meaningful_registry[0][6], "РњРѕР№РєР° Р±Р°Р·РѕРІР°СЏ")
        self.assertEqual(meaningful_registry[0][7], "РњРѕР№РєР°")

        worker_rows = list(workbook["РЎРѕС‚СЂСѓРґРЅРёРєРё"].iter_rows(min_row=2, values_only=True))
        meaningful_workers = [row for row in worker_rows if row[0] != "РќРµС‚ РґР°РЅРЅС‹С…"]
        self.assertEqual(len(meaningful_workers), 1)
        self.assertEqual(meaningful_workers[0][0], "РћР»РµРі")

    def test_admin_create_booking_can_assign_workers_and_notify_them(self) -> None:
        from app.database import SessionLocal
        from app.models import Notification, StaffUser

        admin_token = self.login_staff("admin", "admin")
        with SessionLocal() as db:
            worker = db.scalar(select(StaffUser).where(StaffUser.id == "w1"))
            self.assertIsNotNone(worker)
            assert worker is not None
            worker.telegram_chat_id = "555777999"
            db.commit()

        sent_messages: list[tuple[str | int, str]] = []

        def fake_send_message(chat_id: str | int, text: str) -> None:
            sent_messages.append((chat_id, text))

        with patch("app.main.send_telegram_message", side_effect=fake_send_message):
            response = self.client.post(
                "/api/bookings",
                headers=self.auth_headers(admin_token),
                json={
                    "clientId": "",
                    "clientName": "Pavel",
                    "clientPhone": "+7 (999) 222-33-44",
                    "service": "РњРѕР№РєР° Р±Р°Р·РѕРІР°СЏ",
                    "serviceId": "s1",
                    "date": self.next_active_date(),
                    "time": "12:00",
                    "duration": 30,
                    "price": 1200,
                    "status": "scheduled",
                    "workers": [{"workerId": "w1", "workerName": "РРІР°РЅ", "percent": 35}],
                    "box": "Р‘РѕРєСЃ 1",
                    "paymentType": "cash",
                    "car": "Lada Vesta",
                    "plate": "A123BC",
                    "notes": "urgent wash",
                    "notifyWorkers": True,
                },
            )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(len(payload["workers"]), 1)
        self.assertEqual(payload["workers"][0]["workerId"], "w1")
        self.assertEqual(payload["workers"][0]["percent"], 35)

        self.assertEqual(len(sent_messages), 1)
        self.assertEqual(str(sent_messages[0][0]), "555777999")
        self.assertIn("Pavel", sent_messages[0][1])
        self.assertIn("35%", sent_messages[0][1])
        self.assertIn("urgent wash", sent_messages[0][1])

        with SessionLocal() as db:
            worker_notifications = db.scalars(
                select(Notification).where(
                    Notification.recipient_role == "worker",
                    Notification.recipient_id == "w1",
                )
            ).all()
        self.assertEqual(len(worker_notifications), 1)
        self.assertIn("35%", worker_notifications[0].message)
        self.assertIn("urgent wash", worker_notifications[0].message)

    def test_fired_worker_loses_access_and_future_assignments(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        self.disable_owner_two_factor()
        owner_token = self.login_staff("owner", "owner")
        worker = self.get_staff(login="ivan")

        create_response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(admin_token),
            json={
                "clientId": "",
                "clientName": "Alice",
                "clientPhone": "+7 (999) 111-22-33",
                "service": "Р‘Р°Р·РѕРІР°СЏ РјРѕР№РєР°",
                "serviceId": "s1",
                "date": self.next_active_date(),
                "time": "12:00",
                "duration": 30,
                "price": 1200,
                "status": "scheduled",
                "workers": [{"workerId": worker["id"], "workerName": "РРІР°РЅ", "percent": 40}],
                "box": "Р‘РѕРєСЃ 1",
                "paymentType": "cash",
                "car": "Lada Vesta",
                "plate": "A123BC",
            },
        )
        self.assertEqual(create_response.status_code, 200, create_response.text)
        booking_id = create_response.json()["id"]
        self.assertEqual(len(create_response.json()["workers"]), 1)

        worker_token = self.login_staff("ivan", "master")
        dismiss_response = self.client.delete(
            f"/api/workers/{worker['id']}",
            headers=self.auth_headers(owner_token),
        )
        self.assertEqual(dismiss_response.status_code, 200, dismiss_response.text)

        session_response = self.client.get("/api/auth/session", headers=self.auth_headers(worker_token))
        self.assertEqual(session_response.status_code, 401, session_response.text)

        relogin_response = self.client.post(
            "/api/auth/staff/login",
            json={"login": "ivan", "password": "master"},
        )
        self.assertEqual(relogin_response.status_code, 401, relogin_response.text)

        owner_bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(owner_token))
        self.assertEqual(owner_bootstrap.status_code, 200, owner_bootstrap.text)
        booking = next(item for item in owner_bootstrap.json()["bookings"] if item["id"] == booking_id)
        self.assertEqual(booking["workers"], [])

        fired_worker = self.get_staff(staff_id=worker["id"])
        self.assertEqual(fired_worker["role"], "dismissed_worker")
        self.assertFalse(bool(fired_worker["active"]))

    def test_same_telegram_client_reuses_existing_account(self) -> None:
        first = self.client.post(
            "/api/auth/client",
            json=self.client_auth_payload(name="Alice", phone="+7 (999) 111-22-33", telegram_id="1001"),
        )
        self.assertEqual(first.status_code, 200, first.text)
        first_payload = first.json()

        second = self.client.post(
            "/api/auth/client",
            json=self.client_auth_payload(
                name="Alice Updated",
                phone="8 (999) 111-22-33",
                car="Kia Rio",
                plate="B222BB",
                telegram_id="1001",
            ),
        )
        self.assertEqual(second.status_code, 200, second.text)
        second_payload = second.json()

        self.assertEqual(second_payload["actorId"], first_payload["actorId"])
        self.assertEqual(self.count_clients(), 1)
        client = self.get_client(first_payload["actorId"])
        self.assertEqual(client["phone"], "+7 (999) 111-22-33")
        self.assertEqual(client["telegram_id"], "1001")
        self.assertEqual(client["car"], "Kia Rio")
        self.assertEqual(client["plate"], "B222BB")

    def test_generic_telegram_auth_logs_in_linked_client(self) -> None:
        first = self.client.post(
            "/api/auth/client",
            json=self.client_auth_payload(name="Alice", phone="+7 (999) 111-22-33", telegram_id="1001"),
        )
        self.assertEqual(first.status_code, 200, first.text)

        response = self.client.post("/api/auth/telegram", json={"initData": self.make_init_data("1001")})
        self.assertEqual(response.status_code, 200, response.text)

        payload = response.json()
        self.assertEqual(payload["role"], "client")
        self.assertEqual(payload["actorId"], first.json()["actorId"])
        self.assertEqual(payload["bootstrap"]["session"]["role"], "client")
        self.assertEqual(payload["bootstrap"]["clientProfile"]["name"], "Alice")

    def test_generic_telegram_auth_tolerates_legacy_client_profile_data(self) -> None:
        from app.database import SessionLocal
        from app.models import Client

        first = self.client.post(
            "/api/auth/client",
            json=self.client_auth_payload(name="Alice", phone="+7 (999) 111-22-33", telegram_id="1001"),
        )
        self.assertEqual(first.status_code, 200, first.text)
        client_id = first.json()["actorId"]

        with SessionLocal() as db:
            client = db.get(Client, client_id)
            self.assertIsNotNone(client)
            assert client is not None
            client.name = "1"
            client.phone = "broken-phone"
            client.car = "***"
            client.plate = "###"
            db.commit()

        response = self.client.post("/api/auth/telegram", json={"initData": self.make_init_data("1001")})
        self.assertEqual(response.status_code, 200, response.text)

        payload = response.json()
        self.assertEqual(payload["role"], "client")
        self.assertEqual(payload["actorId"], client_id)
        self.assertEqual(payload["bootstrap"]["clientProfile"]["phone"], "broken-phone")

    def test_generic_telegram_auth_prefers_linked_staff_window(self) -> None:
        self.set_staff_telegram("ivan", "7001")
        client = self.client.post(
            "/api/auth/client",
            json=self.client_auth_payload(name="Alice", phone="+7 (999) 111-22-33", telegram_id="7001"),
        )
        self.assertEqual(client.status_code, 200, client.text)

        response = self.client.post("/api/auth/telegram", json={"initData": self.make_init_data("7001", first_name="Ivan")})
        self.assertEqual(response.status_code, 200, response.text)

        payload = response.json()
        self.assertEqual(payload["role"], "worker")
        self.assertEqual(payload["bootstrap"]["session"]["role"], "worker")
        self.assertEqual(payload["actorId"], self.get_staff(login="ivan")["id"])

    def test_generic_telegram_auth_does_not_claim_primary_owner(self) -> None:
        self.set_primary_owner_telegram("")

        response = self.client.post("/api/auth/telegram", json={"initData": self.make_init_data("9001", first_name="Owner")})
        self.assertEqual(response.status_code, 404, response.text)
        self.assertEqual(self.get_staff(login="owner")["telegram_chat_id"], "")

    def test_primary_owner_telegram_route_rejects_unlinked_owner(self) -> None:
        self.set_primary_owner_telegram("")

        response = self.client.post(
            "/api/auth/telegram-owner",
            json={"initData": self.make_init_data("9001", first_name="Owner")},
        )
        self.assertEqual(response.status_code, 409, response.text)
        self.assertEqual(self.get_staff(login="owner")["telegram_chat_id"], "")

    def test_primary_owner_can_log_in_via_dedicated_telegram_route(self) -> None:
        self.set_primary_owner_telegram("9001")

        response = self.client.post(
            "/api/auth/telegram-owner",
            json={"initData": self.make_init_data("9001", first_name="Owner")},
        )
        self.assertEqual(response.status_code, 200, response.text)

        payload = response.json()
        self.assertEqual(payload["role"], "owner")
        self.assertEqual(payload["bootstrap"]["session"]["role"], "owner")
        self.assertEqual(payload["actorId"], self.get_staff(login="owner")["id"])
        self.assertEqual(self.get_staff(login="owner")["telegram_chat_id"], "9001")

    def test_generic_telegram_auth_rejects_expired_init_data(self) -> None:
        self.set_staff_telegram("ivan", "7002")

        response = self.client.post(
            "/api/auth/telegram",
            json={"initData": self.make_init_data("7002", first_name="Ivan", auth_date=int(time.time()) - 172800)},
        )
        self.assertEqual(response.status_code, 401, response.text)
        self.assertIn("expired", response.text.lower())

    def test_generic_telegram_auth_rejects_duplicate_staff_bindings(self) -> None:
        self.set_staff_telegram("ivan", "7007")
        self.set_staff_telegram("oleg", "7007")

        response = self.client.post("/api/auth/telegram", json={"initData": self.make_init_data("7007", first_name="Ivan")})
        self.assertEqual(response.status_code, 409, response.text)
        self.assertIn("telegram", response.text.lower())

    def test_client_registration_rejects_same_phone_for_different_telegram_ids(self) -> None:
        first = self.client.post(
            "/api/auth/client",
            json=self.client_auth_payload(name="Alice", phone="+7 (999) 111-22-33", telegram_id="1001"),
        )
        self.assertEqual(first.status_code, 200, first.text)

        second = self.client.post(
            "/api/auth/client",
            json=self.client_auth_payload(name="Bob", phone="8 (999) 111-22-33", telegram_id="2002"),
        )
        self.assertEqual(second.status_code, 409, second.text)
        self.assertEqual(self.count_clients(), 1)

    def test_client_profile_cannot_take_phone_of_another_client(self) -> None:
        first_token, first_id = self.login_client(name="Alice", phone="+7 (999) 111-22-33")
        second_token, _ = self.login_client(name="Bob", phone="+7 (999) 222-33-44")

        response = self.client.patch(
            "/api/clients/me",
            headers=self.auth_headers(second_token),
            json={
                "name": "Bob",
                "phone": "8 (999) 111-22-33",
                "car": "Lada Granta",
                "plate": "C333CC",
                "registered": True,
            },
        )
        self.assertEqual(response.status_code, 409, response.text)
        first_client = self.get_client(first_id)
        self.assertEqual(first_client["phone"], "+7 (999) 111-22-33")

    def test_client_booking_creates_notification_for_same_client_id(self) -> None:
        token, actor_id = self.login_client(name="Alice", phone="+7 (999) 111-22-33")
        response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(token),
            json={
                "clientId": "",
                "clientName": "Alice",
                "clientPhone": "+7 (999) 111-22-33",
                "service": "??????? ?????",
                "serviceId": "s1",
                "date": self.next_active_date(),
                "time": "10:00",
                "duration": 30,
                "price": 1200,
                "status": "scheduled",
                "workers": [],
                "box": "???? 1",
                "paymentType": "cash",
                "car": "Lada Vesta",
                "plate": "A123BC",
            },
        )
        self.assertEqual(response.status_code, 200, response.text)

        bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(token))
        self.assertEqual(bootstrap.status_code, 200, bootstrap.text)
        notifications = bootstrap.json()["notifications"]
        self.assertTrue(notifications)
        self.assertTrue(any(item["recipientId"] == actor_id and item["recipientRole"] == "client" for item in notifications))

    def test_client_cannot_mark_other_clients_notification_as_read(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        first_token, first_id = self.login_client(name="Alice", phone="+7 (999) 111-22-33")
        second_token, _ = self.login_client(name="Bob", phone="+7 (999) 222-33-44")

        notification = self.client.post(
            "/api/notifications",
            headers=self.auth_headers(admin_token),
            json={
                "recipientRole": "client",
                "recipientId": first_id,
                "message": "?????? ??? Alice",
                "read": False,
            },
        )
        self.assertEqual(notification.status_code, 200, notification.text)
        notification_id = notification.json()["id"]

        foreign_read = self.client.patch(
            f"/api/notifications/{notification_id}/read",
            headers=self.auth_headers(second_token),
        )
        self.assertEqual(foreign_read.status_code, 403, foreign_read.text)

        own_bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(first_token))
        self.assertEqual(own_bootstrap.status_code, 200, own_bootstrap.text)
        self.assertTrue(any(item["id"] == notification_id and item["read"] is False for item in own_bootstrap.json()["notifications"]))


    def test_client_login_rejects_foreign_telegram_id_for_existing_phone(self) -> None:
        first = self.client.post(
            "/api/auth/client",
            json=self.client_auth_payload(name="Alice", phone="+7 (999) 111-22-33", telegram_id="1001"),
        )
        self.assertEqual(first.status_code, 200, first.text)

        insecure = self.client.post(
            "/api/auth/client",
            json=self.client_auth_payload(name="Bob", phone="+7 (999) 222-33-44"),
        )
        self.assertEqual(insecure.status_code, 200, insecure.text)

        relink_attempt = self.client.post(
            "/api/auth/client",
            json=self.client_auth_payload(name="Bob", phone="+7 (999) 222-33-44", telegram_id="1001"),
        )
        self.assertEqual(relink_attempt.status_code, 409, relink_attempt.text)
        self.assertEqual(self.count_clients(), 2)

    def test_client_read_all_marks_only_own_notifications(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        first_token, first_id = self.login_client(name="Alice", phone="+7 (999) 111-22-33")
        second_token, second_id = self.login_client(name="Bob", phone="+7 (999) 222-33-44")

        first_notification = self.client.post(
            "/api/notifications",
            headers=self.auth_headers(admin_token),
            json={
                "recipientRole": "client",
                "recipientId": first_id,
                "message": "first",
                "read": False,
            },
        )
        self.assertEqual(first_notification.status_code, 200, first_notification.text)
        second_notification = self.client.post(
            "/api/notifications",
            headers=self.auth_headers(admin_token),
            json={
                "recipientRole": "client",
                "recipientId": second_id,
                "message": "second",
                "read": False,
            },
        )
        self.assertEqual(second_notification.status_code, 200, second_notification.text)

        read_all = self.client.post(
            "/api/notifications/read-all",
            headers=self.auth_headers(first_token),
            json={"role": "client"},
        )
        self.assertEqual(read_all.status_code, 200, read_all.text)

        first_bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(first_token))
        second_bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(second_token))
        self.assertEqual(first_bootstrap.status_code, 200, first_bootstrap.text)
        self.assertEqual(second_bootstrap.status_code, 200, second_bootstrap.text)
        self.assertTrue(all(item["read"] is True for item in first_bootstrap.json()["notifications"]))
        self.assertTrue(any(item["read"] is False for item in second_bootstrap.json()["notifications"]))

    def test_client_read_all_rejects_foreign_role_payload(self) -> None:
        token, _ = self.login_client(name="Alice", phone="+7 (999) 111-22-33")
        response = self.client.post(
            "/api/notifications/read-all",
            headers=self.auth_headers(token),
            json={"role": "admin"},
        )
        self.assertEqual(response.status_code, 403, response.text)

    def test_deleting_client_removes_client_sessions_and_notifications(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        client_token, client_id = self.login_client(name="Alice", phone="+7 (999) 111-22-33")

        notification = self.client.post(
            "/api/notifications",
            headers=self.auth_headers(admin_token),
            json={
                "recipientRole": "client",
                "recipientId": client_id,
                "message": "cleanup",
                "read": False,
            },
        )
        self.assertEqual(notification.status_code, 200, notification.text)
        self.assertGreaterEqual(self.count_client_notifications(client_id), 1)
        self.assertGreaterEqual(self.count_client_sessions(client_id), 1)

        delete_response = self.client.delete(
            f"/api/clients/{client_id}",
            headers=self.auth_headers(admin_token),
        )
        self.assertEqual(delete_response.status_code, 200, delete_response.text)
        self.assertEqual(self.count_clients(), 0)
        self.assertEqual(self.count_client_notifications(client_id), 0)
        self.assertEqual(self.count_client_sessions(client_id), 0)

        session_response = self.client.get("/api/auth/session", headers=self.auth_headers(client_token))
        self.assertEqual(session_response.status_code, 401, session_response.text)

    def test_client_cancel_booking_creates_client_and_admin_notifications(self) -> None:
        token, actor_id = self.login_client(name="Alice", phone="+7 (999) 111-22-33")
        create_response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(token),
            json={
                "clientId": "",
                "clientName": "Alice",
                "clientPhone": "+7 (999) 111-22-33",
                "service": "??????? ?????",
                "serviceId": "s1",
                "date": self.next_active_date(),
                "time": "11:00",
                "duration": 30,
                "price": 1200,
                "status": "scheduled",
                "workers": [],
                "box": "???? 1",
                "paymentType": "cash",
                "car": "Lada Vesta",
                "plate": "A123BC",
            },
        )
        self.assertEqual(create_response.status_code, 200, create_response.text)
        booking_id = create_response.json()["id"]

        delete_response = self.client.delete(
            f"/api/bookings/{booking_id}",
            headers=self.auth_headers(token),
        )
        self.assertEqual(delete_response.status_code, 200, delete_response.text)

        client_bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(token))
        admin_token = self.login_staff("admin", "admin")
        admin_bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(admin_token))
        self.assertEqual(client_bootstrap.status_code, 200, client_bootstrap.text)
        self.assertEqual(admin_bootstrap.status_code, 200, admin_bootstrap.text)
        self.assertTrue(any(item["recipientId"] == actor_id and item["recipientRole"] == "client" for item in client_bootstrap.json()["notifications"]))
        self.assertTrue(any(item["recipientRole"] == "admin" for item in admin_bootstrap.json()["notifications"]))


    def test_deleted_client_can_register_again_with_same_phone_and_telegram(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        first = self.client.post(
            "/api/auth/client",
            json=self.client_auth_payload(name="Alice", phone="+7 (999) 111-22-33", telegram_id="1001"),
        )
        self.assertEqual(first.status_code, 200, first.text)
        first_id = first.json()["actorId"]

        delete_response = self.client.delete(
            f"/api/clients/{first_id}",
            headers=self.auth_headers(admin_token),
        )
        self.assertEqual(delete_response.status_code, 200, delete_response.text)

        second = self.client.post(
            "/api/auth/client",
            json=self.client_auth_payload(name="Alice", phone="8 (999) 111-22-33", telegram_id="1001"),
        )
        self.assertEqual(second.status_code, 200, second.text)
        self.assertNotEqual(second.json()["actorId"], first_id)
        self.assertEqual(self.count_clients(), 1)

    def test_secure_client_auth_requires_valid_init_data(self) -> None:
        self.shutdown_app()
        os.environ["ALLOW_INSECURE_CLIENT_AUTH"] = "false"
        self.restart_app()

        missing = self.client.post(
            "/api/auth/client",
            json=self.client_auth_payload(name="Alice", phone="+7 (999) 111-22-33"),
        )
        self.assertEqual(missing.status_code, 401, missing.text)

        invalid = self.client.post(
            "/api/auth/client",
            json={
                **self.client_auth_payload(name="Alice", phone="+7 (999) 111-22-33"),
                "initData": "user=%7B%22id%22%3A1001%7D&hash=bad",
            },
        )
        self.assertEqual(invalid.status_code, 401, invalid.text)

        valid = self.client.post(
            "/api/auth/client",
            json=self.client_auth_payload(name="Alice", phone="+7 (999) 111-22-33", telegram_id="1001"),
        )
        self.assertEqual(valid.status_code, 200, valid.text)

        self.shutdown_app()
        os.environ["ALLOW_INSECURE_CLIENT_AUTH"] = "true"
        self.restart_app()

    def test_admin_reschedule_creates_client_notification(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        client_token, actor_id = self.login_client(name="Alice", phone="+7 (999) 111-22-33")
        create_response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(client_token),
            json={
                "clientId": "",
                "clientName": "Alice",
                "clientPhone": "+7 (999) 111-22-33",
                "service": "??????? ?????",
                "serviceId": "s1",
                "date": self.next_active_date(),
                "time": "13:00",
                "duration": 30,
                "price": 1200,
                "status": "scheduled",
                "workers": [],
                "box": "???? 1",
                "paymentType": "cash",
                "car": "Lada Vesta",
                "plate": "A123BC",
            },
        )
        self.assertEqual(create_response.status_code, 200, create_response.text)
        booking_id = create_response.json()["id"]

        update_response = self.client.patch(
            f"/api/bookings/{booking_id}",
            headers=self.auth_headers(admin_token),
            json={"time": "14:00"},
        )
        self.assertEqual(update_response.status_code, 200, update_response.text)

        client_bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(client_token))
        self.assertEqual(client_bootstrap.status_code, 200, client_bootstrap.text)
        self.assertTrue(any(item["recipientId"] == actor_id and "14:00" in item["message"] for item in client_bootstrap.json()["notifications"]))

    def test_admin_completion_creates_client_notification(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        client_token, actor_id = self.login_client(name="Alice", phone="+7 (999) 111-22-33")
        create_response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(client_token),
            json={
                "clientId": "",
                "clientName": "Alice",
                "clientPhone": "+7 (999) 111-22-33",
                "service": "??????? ?????",
                "serviceId": "s1",
                "date": self.next_active_date(),
                "time": "15:00",
                "duration": 30,
                "price": 1200,
                "status": "scheduled",
                "workers": [],
                "box": "???? 1",
                "paymentType": "cash",
                "car": "Lada Vesta",
                "plate": "A123BC",
            },
        )
        self.assertEqual(create_response.status_code, 200, create_response.text)
        booking_id = create_response.json()["id"]

        update_response = self.client.patch(
            f"/api/bookings/{booking_id}",
            headers=self.auth_headers(admin_token),
            json={"status": "completed"},
        )
        self.assertEqual(update_response.status_code, 200, update_response.text)

        client_bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(client_token))
        self.assertEqual(client_bootstrap.status_code, 200, client_bootstrap.text)
        self.assertTrue(any(item["recipientId"] == actor_id and "?????????" in item["message"].lower() for item in client_bootstrap.json()["notifications"]))


    def test_admin_booking_reuses_existing_client_by_normalized_phone(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        _, client_id = self.login_client(name="Alice", phone="+7 (999) 111-22-33")

        response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(admin_token),
            json={
                "clientId": "",
                "clientName": "Alice Admin",
                "clientPhone": "8 (999) 111-22-33",
                "service": "??????? ?????",
                "serviceId": "s1",
                "date": self.next_active_date(),
                "time": "16:00",
                "duration": 30,
                "price": 1200,
                "status": "scheduled",
                "workers": [],
                "box": "???? 1",
                "paymentType": "cash",
                "car": "Kia Rio",
                "plate": "B222BB",
            },
        )
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["clientId"], client_id)
        self.assertEqual(self.count_clients(), 1)
        client = self.get_client(client_id)
        self.assertEqual(client["phone"], "+7 (999) 111-22-33")
        self.assertEqual(client["name"], "Alice Admin")
        self.assertEqual(client["car"], "Kia Rio")

    def test_admin_cannot_create_booking_with_conflicting_client_and_phone(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        _, first_client_id = self.login_client(name="Alice", phone="+7 (999) 111-22-33")
        self.login_client(name="Bob", phone="+7 (999) 222-33-44")

        response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(admin_token),
            json={
                "clientId": first_client_id,
                "clientName": "Alice",
                "clientPhone": "+7 (999) 222-33-44",
                "service": "??????? ?????",
                "serviceId": "s1",
                "date": self.next_active_date(),
                "time": "17:00",
                "duration": 30,
                "price": 1200,
                "status": "scheduled",
                "workers": [],
                "box": "???? 1",
                "paymentType": "cash",
                "car": "Lada Vesta",
                "plate": "A123BC",
            },
        )
        self.assertEqual(response.status_code, 409, response.text)

    def test_admin_can_save_profile_and_notification_settings(self) -> None:
        admin_token = self.login_staff("admin", "admin")

        profile_response = self.client.put(
            "/api/settings/admin/profile",
            headers=self.auth_headers(admin_token),
            json={
                "name": "??????? ?????",
                "email": "admin@example.com",
                "phone": "+7 (912) 000-11-22",
                "telegramChatId": "555123",
            },
        )
        self.assertEqual(profile_response.status_code, 200, profile_response.text)

        notif_response = self.client.put(
            "/api/settings/admin/notifications",
            headers=self.auth_headers(admin_token),
            json={
                "newBooking": True,
                "cancelled": False,
                "paymentDue": True,
                "workerAssigned": False,
                "reminders": True,
            },
        )
        self.assertEqual(notif_response.status_code, 200, notif_response.text)

        bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(admin_token))
        self.assertEqual(bootstrap.status_code, 200, bootstrap.text)
        settings = bootstrap.json()["settings"]
        self.assertEqual(settings["adminProfile"]["name"], "??????? ?????")
        self.assertEqual(settings["adminProfile"]["telegramChatId"], "555123")
        self.assertEqual(settings["adminNotificationSettings"]["paymentDue"], True)
        self.assertEqual(settings["adminNotificationSettings"]["workerAssigned"], False)

    def test_owner_can_create_admin_like_worker_and_update_telegram_ids(self) -> None:
        self.disable_owner_two_factor()
        owner_token = self.login_staff("owner", "owner")

        create_admin = self.client.post(
            "/api/workers",
            headers=self.auth_headers(owner_token),
            json={
                "role": "admin",
                "name": "Shift Admin",
                "login": "shiftadmin",
                "password": "adminpass",
                "percent": 0,
                "salaryBase": 45000,
                "phone": "+7 (999) 555-11-22",
                "email": "shiftadmin@example.com",
                "telegramChatId": "701001",
            },
        )
        self.assertEqual(create_admin.status_code, 200, create_admin.text)
        admin_payload = create_admin.json()
        self.assertEqual(admin_payload["role"], "admin")
        self.assertEqual(admin_payload["telegramChatId"], "701001")

        create_worker = self.client.post(
            "/api/workers",
            headers=self.auth_headers(owner_token),
            json={
                "role": "worker",
                "name": "Detail Master",
                "login": "detailmaster",
                "password": "workerpass",
                "percent": 35,
                "salaryBase": 15000,
                "phone": "+7 (999) 555-22-33",
                "email": "detailmaster@example.com",
                "telegramChatId": "701002",
            },
        )
        self.assertEqual(create_worker.status_code, 200, create_worker.text)
        worker_payload = create_worker.json()
        self.assertEqual(worker_payload["role"], "worker")
        self.assertEqual(worker_payload["telegramChatId"], "701002")

        update_settings = self.client.put(
            "/api/workers/settings",
            headers=self.auth_headers(owner_token),
            json=[
                {
                    "id": admin_payload["id"],
                    "role": "admin",
                    "name": "Shift Admin",
                    "percent": 0,
                    "salaryBase": 50000,
                    "active": True,
                    "telegramChatId": "801001",
                },
                {
                    "id": worker_payload["id"],
                    "role": "worker",
                    "name": "Detail Master",
                    "percent": 30,
                    "salaryBase": 18000,
                    "active": True,
                    "telegramChatId": "801002",
                },
            ],
        )
        self.assertEqual(update_settings.status_code, 200, update_settings.text)
        saved_staff = {item["id"]: item for item in update_settings.json()}
        self.assertEqual(saved_staff[admin_payload["id"]]["telegramChatId"], "801001")
        self.assertEqual(saved_staff[admin_payload["id"]]["role"], "admin")
        self.assertEqual(saved_staff[worker_payload["id"]]["telegramChatId"], "801002")
        self.assertEqual(saved_staff[worker_payload["id"]]["role"], "worker")

        admin_record = self.get_staff(login="shiftadmin")
        worker_record = self.get_staff(login="detailmaster")
        self.assertEqual(admin_record["role"], "admin")
        self.assertEqual(admin_record["telegram_chat_id"], "801001")
        self.assertEqual(worker_record["role"], "worker")
        self.assertEqual(worker_record["telegram_chat_id"], "801002")

    def test_owner_can_create_and_login_accountant(self) -> None:
        self.disable_owner_two_factor()
        owner_token = self.login_staff("owner", "owner")

        create_accountant = self.client.post(
            "/api/workers",
            headers=self.auth_headers(owner_token),
            json={
                "role": "accountant",
                "name": "Finance Lead",
                "login": "accountant",
                "password": "accpass",
                "percent": 0,
                "salaryBase": 60000,
                "phone": "+7 (999) 555-33-44",
                "email": "accountant@example.com",
                "telegramChatId": "701003",
            },
        )
        self.assertEqual(create_accountant.status_code, 200, create_accountant.text)
        accountant_payload = create_accountant.json()
        self.assertEqual(accountant_payload["role"], "accountant")
        self.assertEqual(accountant_payload["telegramChatId"], "701003")

        accountant_login = self.client.post(
            "/api/auth/staff/login",
            json={"login": "accountant", "password": "accpass"},
        )
        self.assertEqual(accountant_login.status_code, 200, accountant_login.text)
        bootstrap = accountant_login.json()["bootstrap"]
        self.assertEqual(bootstrap["session"]["role"], "accountant")
        self.assertTrue(any(item["id"] == accountant_payload["id"] for item in bootstrap["workers"]))

        bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(owner_token))
        self.assertEqual(bootstrap.status_code, 200, bootstrap.text)
        owner_workers = {item["id"]: item for item in bootstrap.json()["workers"]}
        self.assertEqual(owner_workers[accountant_payload["id"]]["role"], "accountant")
        self.assertEqual(owner_workers[accountant_payload["id"]]["telegramChatId"], "701003")

    def test_owner_can_rehire_employee_with_same_telegram_after_dismissal(self) -> None:
        self.disable_owner_two_factor()
        owner_token = self.login_staff("owner", "owner")

        create_worker = self.client.post(
            "/api/workers",
            headers=self.auth_headers(owner_token),
            json={
                "role": "worker",
                "name": "Repeat Master",
                "login": "repeatmaster",
                "password": "workerpass",
                "percent": 35,
                "salaryBase": 15000,
                "phone": "+7 (999) 555-77-88",
                "email": "repeatmaster@example.com",
                "telegramChatId": "909001",
            },
        )
        self.assertEqual(create_worker.status_code, 200, create_worker.text)
        first_worker = create_worker.json()

        dismiss_response = self.client.delete(
            f"/api/workers/{first_worker['id']}",
            headers=self.auth_headers(owner_token),
        )
        self.assertEqual(dismiss_response.status_code, 200, dismiss_response.text)

        dismissed_record = self.get_staff(staff_id=first_worker["id"])
        self.assertEqual(dismissed_record["role"], "dismissed_worker")
        self.assertEqual(dismissed_record["telegram_chat_id"], "")

        rehire_response = self.client.post(
            "/api/workers",
            headers=self.auth_headers(owner_token),
            json={
                "role": "worker",
                "name": "Repeat Master 2",
                "login": "repeatmaster2",
                "password": "workerpass2",
                "percent": 30,
                "salaryBase": 18000,
                "phone": "+7 (999) 555-88-99",
                "email": "repeatmaster2@example.com",
                "telegramChatId": "909001",
            },
        )
        self.assertEqual(rehire_response.status_code, 200, rehire_response.text)
        rehired_worker = rehire_response.json()
        self.assertEqual(rehired_worker["telegramChatId"], "909001")

    def test_admin_can_manage_master_payroll_and_private_client_rating(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        _, client_id = self.login_client(name="Alice", phone="+7 (999) 111-22-33")

        payroll_response = self.client.put(
            "/api/admin/workers/payroll",
            headers=self.auth_headers(admin_token),
            json=[
                {
                    "id": "w1",
                    "role": "worker",
                    "name": "Иван",
                    "percent": 28,
                    "salaryBase": 22000,
                    "active": True,
                    "telegramChatId": "",
                },
            ],
        )
        self.assertEqual(payroll_response.status_code, 200, payroll_response.text)
        payroll_workers = {item["id"]: item for item in payroll_response.json()}
        self.assertEqual(payroll_workers["w1"]["defaultPercent"], 28)
        self.assertEqual(payroll_workers["w1"]["salaryBase"], 22000)

        card_response = self.client.patch(
            f"/api/clients/{client_id}/card",
            headers=self.auth_headers(admin_token),
            json={
                "adminRating": 4,
                "adminNote": "Нужен звонок перед подтверждением",
            },
        )
        self.assertEqual(card_response.status_code, 200, card_response.text)
        self.assertEqual(card_response.json()["adminRating"], 4)
        self.assertEqual(card_response.json()["adminNote"], "Нужен звонок перед подтверждением")

        bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(admin_token))
        self.assertEqual(bootstrap.status_code, 200, bootstrap.text)
        clients = {item["id"]: item for item in bootstrap.json()["clients"]}
        self.assertEqual(clients[client_id]["adminRating"], 4)
        self.assertEqual(clients[client_id]["adminNote"], "Нужен звонок перед подтверждением")

    def test_owner_and_admin_can_see_detailed_worker_payroll_summary(self) -> None:
        self.disable_owner_two_factor()
        owner_token = self.login_staff("owner", "owner")
        admin_token = self.login_staff("admin", "admin")

        payroll_response = self.client.put(
            "/api/admin/workers/payroll",
            headers=self.auth_headers(admin_token),
            json=[
                {
                    "id": "w1",
                    "role": "worker",
                    "name": "Иван",
                    "percent": 20,
                    "salaryBase": 1000,
                    "active": True,
                    "telegramChatId": "",
                },
            ],
        )
        self.assertEqual(payroll_response.status_code, 200, payroll_response.text)

        booking_response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(admin_token),
            json={
                "clientId": "",
                "clientName": "Payroll Client",
                "clientPhone": "+7 (999) 222-33-44",
                "service": "Мойка базовая",
                "serviceId": "s1",
                "date": self.next_active_date(),
                "time": "12:00",
                "duration": 30,
                "price": 5000,
                "status": "completed",
                "workers": [{"workerId": "w1", "workerName": "Иван", "percent": 20}],
                "box": "Бокс 1",
                "paymentType": "cash",
                "car": "Lada Vesta",
                "plate": "A123BC",
                "notifyWorkers": False,
            },
        )
        self.assertEqual(booking_response.status_code, 200, booking_response.text)

        advance_response = self.client.post(
            "/api/payroll/entries",
            headers=self.auth_headers(admin_token),
            json={
                "workerId": "w1",
                "kind": "advance",
                "amount": 300,
                "note": "Аванс на материалы",
            },
        )
        self.assertEqual(advance_response.status_code, 200, advance_response.text)

        bonus_response = self.client.post(
            "/api/payroll/entries",
            headers=self.auth_headers(owner_token),
            json={
                "workerId": "w1",
                "kind": "bonus",
                "amount": 500,
                "note": "Премия за допродажу",
            },
        )
        self.assertEqual(bonus_response.status_code, 200, bonus_response.text)

        owner_bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(owner_token))
        self.assertEqual(owner_bootstrap.status_code, 200, owner_bootstrap.text)
        owner_worker = {item["id"]: item for item in owner_bootstrap.json()["workers"]}["w1"]
        owner_summary = owner_worker["payrollSummary"]
        self.assertEqual(owner_summary["completedBookings"], 1)
        self.assertEqual(owner_summary["completedRevenue"], 5000)
        self.assertEqual(owner_summary["accruedFromBookings"], 1000)
        self.assertEqual(owner_summary["baseSalary"], 1000)
        self.assertEqual(owner_summary["bonusTotal"], 500)
        self.assertEqual(owner_summary["advanceTotal"], 300)
        self.assertEqual(owner_summary["totalAccrued"], 2500)
        self.assertEqual(owner_summary["totalDeducted"], 300)
        self.assertEqual(owner_summary["balance"], 2200)
        self.assertEqual(owner_summary["bookingItems"][0]["service"], "Мойка базовая")
        self.assertEqual({item["kind"] for item in owner_summary["entries"]}, {"advance", "bonus"})

        admin_bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(admin_token))
        self.assertEqual(admin_bootstrap.status_code, 200, admin_bootstrap.text)
        admin_worker = {item["id"]: item for item in admin_bootstrap.json()["workers"]}["w1"]
        admin_summary = admin_worker["payrollSummary"]
        self.assertEqual(admin_summary["balance"], 2200)
        self.assertEqual(admin_summary["bookingItems"][0]["earned"], 1000)

    def test_payroll_entry_notifies_worker_and_updates_summary(self) -> None:
        from app.database import SessionLocal
        from app.models import Notification, StaffUser

        self.disable_owner_two_factor()
        owner_token = self.login_staff("owner", "owner")

        with SessionLocal() as db:
            worker = db.get(StaffUser, "w1")
            assert worker is not None
            worker.telegram_chat_id = "777888999"
            db.commit()

        sent_messages: list[tuple[str | int, str]] = []

        def fake_send_message(chat_id: str | int, text: str) -> None:
            sent_messages.append((chat_id, text))

        with patch("app.main.send_telegram_message", side_effect=fake_send_message):
            response = self.client.post(
                "/api/payroll/entries",
                headers=self.auth_headers(owner_token),
                json={
                    "workerId": "w1",
                    "kind": "payout",
                    "amount": 1200,
                    "note": "Выплата за неделю",
                },
            )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["id"], "w1")
        self.assertTrue(any(entry["kind"] == "payout" and entry["amount"] == 1200 for entry in payload["payrollSummary"]["entries"]))

        worker_token = self.login_staff("ivan", "master")
        worker_bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(worker_token))
        self.assertEqual(worker_bootstrap.status_code, 200, worker_bootstrap.text)
        worker_payload = worker_bootstrap.json()["staffProfile"]
        self.assertTrue(any(entry["kind"] == "payout" and entry["amount"] == 1200 for entry in worker_payload["payrollSummary"]["entries"]))

    def test_admin_cannot_issue_advance_before_worker_earns_1000(self) -> None:
        admin_token = self.login_staff("admin", "admin")

        response = self.client.post(
            "/api/payroll/entries",
            headers=self.auth_headers(admin_token),
            json={
                "workerId": "w1",
                "kind": "advance",
                "amount": 300,
                "note": "Ранний аванс",
            },
        )
        self.assertEqual(response.status_code, 400, response.text)
        self.assertIn("минимум 1000", response.text)

    def test_owner_pdf_export_returns_pdf_file(self) -> None:
        self.disable_owner_two_factor()
        owner_token = self.login_staff("owner", "owner")

        response = self.client.get("/api/owner/exports/pdf", headers=self.auth_headers(owner_token))
        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.headers["content-type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_owner_can_create_booking_with_assigned_master(self) -> None:
        self.disable_owner_two_factor()
        owner_token = self.login_staff("owner", "owner")

        response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(owner_token),
            json={
                "clientId": "",
                "clientName": "Owner Client",
                "clientPhone": "+7 (999) 555-22-11",
                "service": "Мойка базовая",
                "serviceId": "s1",
                "date": self.next_active_date(),
                "time": "15:00",
                "duration": 30,
                "price": 1200,
                "status": "confirmed",
                "workers": [{"workerId": "w1", "workerName": "Иван", "percent": 30}],
                "box": "Бокс 1",
                "paymentType": "cash",
                "car": "Lada Vesta",
                "plate": "A123BC",
                "notifyWorkers": False,
            },
        )
        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["workers"][0]["workerId"], "w1")

    def test_admin_reschedule_notifies_assigned_worker(self) -> None:
        from app.database import SessionLocal
        from app.models import Notification, StaffUser

        admin_token = self.login_staff("admin", "admin")
        with SessionLocal() as db:
            worker = db.scalar(select(StaffUser).where(StaffUser.id == "w1"))
            self.assertIsNotNone(worker)
            assert worker is not None
            worker.telegram_chat_id = "777888999"
            db.commit()

        create_response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(admin_token),
            json={
                "clientId": "",
                "clientName": "Reschedule Client",
                "clientPhone": "+7 (999) 444-55-66",
                "service": "Мойка базовая",
                "serviceId": "s1",
                "date": self.next_active_date(),
                "time": "11:00",
                "duration": 30,
                "price": 1200,
                "status": "confirmed",
                "workers": [{"workerId": "w1", "workerName": "Иван", "percent": 30}],
                "box": "Бокс 1",
                "paymentType": "cash",
                "car": "Kia Rio",
                "plate": "A111AA",
                "notifyWorkers": False,
            },
        )
        self.assertEqual(create_response.status_code, 200, create_response.text)
        booking_id = create_response.json()["id"]

        sent_messages: list[tuple[str | int, str]] = []

        def fake_send_message(chat_id: str | int, text: str) -> None:
            sent_messages.append((chat_id, text))

        with patch("app.main.send_telegram_message", side_effect=fake_send_message):
            next_day = (datetime.strptime(self.next_active_date(), "%d.%m.%Y") + timedelta(days=1)).strftime("%d.%m.%Y")
            update_response = self.client.patch(
                f"/api/bookings/{booking_id}",
                headers=self.auth_headers(admin_token),
                json={
                    "date": next_day,
                    "time": "14:30",
                    "box": "Бокс 2",
                    "notes": "Перенос по звонку клиента",
                },
            )
        self.assertEqual(update_response.status_code, 200, update_response.text)

        with SessionLocal() as db:
            notifications = db.scalars(
                select(Notification)
                .where(Notification.recipient_role == "worker", Notification.recipient_id == "w1")
                .order_by(Notification.created_at.desc())
            ).all()

        self.assertTrue(any("перенёс вашу запись" in item.message for item in notifications))
        self.assertTrue(any("Было:" in item.message and "Стало:" in item.message for item in notifications))
        self.assertTrue(any(chat_id == "777888999" and "перенёс вашу запись" in text for chat_id, text in sent_messages))

    def test_worker_start_and_completion_notify_owner_and_send_receipt(self) -> None:
        from app.database import SessionLocal
        from app.models import Client, Notification

        self.disable_owner_two_factor()
        self.set_primary_owner_telegram("123123123")
        self.set_staff_telegram("admin", "456456456")
        admin_token = self.login_staff("admin", "admin")
        worker_token = self.login_staff("ivan", "master")
        _client_token, client_id = self.login_client(name="Alice", phone="+7 (999) 111-22-33")

        with SessionLocal() as db:
            client = db.get(Client, client_id)
            self.assertIsNotNone(client)
            assert client is not None
            client.telegram_id = "999888777"
            db.commit()

        create_response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(admin_token),
            json={
                "clientId": client_id,
                "clientName": "Alice",
                "clientPhone": "+7 (999) 111-22-33",
                "service": "Мойка базовая",
                "serviceId": "s1",
                "date": self.next_active_date(),
                "time": "12:00",
                "duration": 30,
                "price": 1500,
                "status": "scheduled",
                "workers": [{"workerId": "w1", "workerName": "Иван", "percent": 30}],
                "box": "Бокс 1",
                "paymentType": "card",
                "car": "Lada Vesta",
                "plate": "A123BC",
                "notifyWorkers": False,
            },
        )
        self.assertEqual(create_response.status_code, 200, create_response.text)
        booking_id = create_response.json()["id"]

        sent_messages: list[tuple[str | int, str]] = []

        def fake_send_message(chat_id: str | int, text: str) -> None:
            sent_messages.append((chat_id, text))

        with patch("app.main.send_telegram_message", side_effect=fake_send_message):
            start_response = self.client.patch(
                f"/api/bookings/{booking_id}",
                headers=self.auth_headers(worker_token),
                json={"status": "in_progress"},
            )
            self.assertEqual(start_response.status_code, 200, start_response.text)

            complete_response = self.client.patch(
                f"/api/bookings/{booking_id}",
                headers=self.auth_headers(worker_token),
                json={
                    "status": "completed",
                    "price": 1800,
                    "paymentSettled": True,
                    "paymentType": "card",
                    "notes": "Сделали полную уборку салона",
                },
            )
            self.assertEqual(complete_response.status_code, 200, complete_response.text)

        with SessionLocal() as db:
            owner_notifications = db.scalars(
                select(Notification).where(Notification.recipient_role == "owner").order_by(Notification.created_at.desc())
            ).all()
            admin_notifications = db.scalars(
                select(Notification).where(Notification.recipient_role == "admin").order_by(Notification.created_at.desc())
            ).all()
            client_notifications = db.scalars(
                select(Notification)
                .where(Notification.recipient_role == "client", Notification.recipient_id == client_id)
                .order_by(Notification.created_at.desc())
            ).all()

        self.assertTrue(any("Мастер начал работу по записи" in item.message for item in owner_notifications))
        self.assertTrue(any("Мастер завершил работу по записи" in item.message for item in owner_notifications))
        self.assertTrue(any("Чек по записи" in item.message for item in owner_notifications))
        self.assertTrue(any("Чек по записи" in item.message for item in admin_notifications))
        self.assertTrue(any("Чек по записи" in item.message for item in client_notifications))
        self.assertTrue(any(chat_id == "123123123" and "Мастер начал работу по записи" in text for chat_id, text in sent_messages))
        self.assertTrue(any(chat_id == "123123123" and "Чек по записи" in text for chat_id, text in sent_messages))
        self.assertTrue(any(chat_id == "456456456" and "Чек по записи" in text for chat_id, text in sent_messages))
        self.assertTrue(any(chat_id == "999888777" and "Чек по записи" in text for chat_id, text in sent_messages))

    def test_client_can_store_multiple_vehicles(self) -> None:
        token, client_id = self.login_client(name="Alice", phone="+7 (999) 111-22-33")

        response = self.client.patch(
            "/api/clients/me",
            headers=self.auth_headers(token),
            json={
                "name": "Alice",
                "phone": "+7 (999) 111-22-33",
                "car": "Lada Vesta",
                "plate": "А123ВС",
                "vehicles": [
                    {"car": "Lada Vesta", "plate": "А123ВС"},
                    {"car": "Kia Rio", "plate": "К456ОР"},
                ],
                "registered": True,
            },
        )
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(len(payload["vehicles"]), 2)
        self.assertEqual(payload["vehicles"][1]["car"], "Kia Rio")

        bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(token))
        self.assertEqual(bootstrap.status_code, 200, bootstrap.text)
        self.assertEqual(len(bootstrap.json()["clientProfile"]["vehicles"]), 2)

        from app.database import SessionLocal
        from app.models import Client

        with SessionLocal() as db:
            client = db.get(Client, client_id)
            self.assertIsNotNone(client)
            assert client is not None
            self.assertEqual(client.car, "Lada Vesta")
            self.assertEqual(client.plate, "А123ВС")

    def test_owner_can_notify_admin_about_inactive_clients(self) -> None:
        from app.database import SessionLocal
        from app.models import Booking, Notification

        self.disable_owner_two_factor()
        self.set_staff_telegram("admin", "456456456")
        owner_token = self.login_staff("owner", "owner")
        _client_token, client_id = self.login_client(name="Alice", phone="+7 (999) 111-22-33")

        old_date = (datetime.now() - timedelta(days=20)).strftime("%d.%m.%Y")
        with SessionLocal() as db:
            booking = Booking(
                id=f"b-{uuid4()}",
                client_id=client_id,
                client_name="Alice",
                client_phone="+7 (999) 111-22-33",
                service="Мойка базовая",
                service_id="s1",
                date=old_date,
                time="12:00",
                duration=30,
                price=1200,
                status="completed",
                box="Бокс 1",
                payment_type="cash",
                notes="",
                car="Lada Vesta",
                plate="А123ВС",
                created_at=datetime.now(timezone.utc) - timedelta(days=20),
            )
            db.add(booking)
            db.commit()

        sent_messages: list[tuple[str | int, str]] = []

        def fake_send_message(chat_id: str | int, text: str) -> None:
            sent_messages.append((chat_id, text))

        with patch("app.main.send_telegram_message", side_effect=fake_send_message):
            response = self.client.post(
                "/api/owner/inactive-clients/remind-admin",
                headers=self.auth_headers(owner_token),
            )
        self.assertEqual(response.status_code, 200, response.text)
        self.assertIn("Админу отправлено напоминание", response.json()["message"])

        with SessionLocal() as db:
            notifications = db.scalars(
                select(Notification).where(Notification.recipient_role == "admin").order_by(Notification.created_at.desc())
            ).all()

        self.assertTrue(any("не были более двух недель" in item.message for item in notifications))
        self.assertTrue(any("Alice" in item.message for item in notifications))
        self.assertTrue(any(chat_id == "456456456" and "Alice" in text for chat_id, text in sent_messages))

    def test_owner_dispatches_return_visit_reminders_to_clients(self) -> None:
        from app.database import SessionLocal
        from app.models import Booking, Notification

        self.disable_owner_two_factor()
        owner_token = self.login_staff("owner", "owner")
        auth_response = self.client.post(
            "/api/auth/client",
            json=self.client_auth_payload(
                name="Alice",
                phone="+7 (999) 111-22-33",
                telegram_id="999888777",
            ),
        )
        self.assertEqual(auth_response.status_code, 200, auth_response.text)
        client_id = auth_response.json()["bootstrap"]["session"]["actorId"]

        old_date = (datetime.now() - timedelta(days=20)).strftime("%d.%m.%Y")
        with SessionLocal() as db:
            booking = Booking(
                id=f"b-{uuid4()}",
                client_id=client_id,
                client_name="Alice",
                client_phone="+7 (999) 111-22-33",
                service="Мойка базовая",
                service_id="s1",
                date=old_date,
                time="12:00",
                duration=30,
                price=1200,
                status="completed",
                box="Бокс 1",
                payment_type="cash",
                notes="",
                car="Lada Vesta",
                plate="А123ВС",
                created_at=datetime.now(timezone.utc) - timedelta(days=20),
            )
            db.add(booking)
            db.commit()

        sent_messages: list[tuple[str | int, str]] = []

        def fake_send_message(chat_id: str | int, text: str) -> None:
            sent_messages.append((chat_id, text))

        with patch("app.main.send_telegram_message", side_effect=fake_send_message):
            response = self.client.post(
                "/api/owner/reminders/dispatch",
                headers=self.auth_headers(owner_token),
                json={"targetDate": self.next_active_date(), "force": True},
            )
            self.assertEqual(response.status_code, 200, response.text)
            self.assertEqual(response.json()["clientReminders"], 1)

            repeated = self.client.post(
                "/api/owner/reminders/dispatch",
                headers=self.auth_headers(owner_token),
                json={"targetDate": self.next_active_date(), "force": True},
            )
            self.assertEqual(repeated.status_code, 200, repeated.text)
            self.assertEqual(repeated.json()["clientReminders"], 0)

        with SessionLocal() as db:
            notifications = db.scalars(
                select(Notification)
                .where(Notification.recipient_role == "client", Notification.recipient_id == client_id)
                .order_by(Notification.created_at.desc())
            ).all()

        self.assertTrue(any("давно не была чистой" in item.message for item in notifications))
        self.assertTrue(any(chat_id == "999888777" and "Пора вернуться на мойку" in text for chat_id, text in sent_messages))

    def test_worker_can_submit_shift_checklists_and_owner_can_review_them(self) -> None:
        from app.database import SessionLocal
        from app.models import Notification

        self.disable_owner_two_factor()
        owner_token = self.login_staff("owner", "owner")
        worker_token = self.login_staff("ivan", "master")

        for payload in (
            {"name": "Активная пена", "qty": 20, "unit": "л", "unitPrice": 450, "category": "Химия"},
            {"name": "Шампунь", "qty": 12, "unit": "л", "unitPrice": 380, "category": "Химия"},
            {"name": "Перчатки", "qty": 30, "unit": "шт", "unitPrice": 25, "category": "Расходники"},
        ):
            response = self.client.post(
                "/api/stock-items",
                headers=self.auth_headers(owner_token),
                json=payload,
            )
            self.assertEqual(response.status_code, 200, response.text)

        start_response = self.client.post(
            "/api/shift-checklists",
            headers=self.auth_headers(worker_token),
            json={
                "phase": "start",
                "note": "Принял склад без замечаний",
                "items": [],
            },
        )
        self.assertEqual(start_response.status_code, 200, start_response.text)
        start_payload = start_response.json()
        self.assertEqual(start_payload["phase"], "start")
        self.assertEqual(len(start_payload["items"]), 2)

        item_updates = [
            {"stockItemId": item["stockItemId"], "actualQty": max(0, item["actualQty"] - 2)}
            for item in start_payload["items"]
        ]
        end_response = self.client.post(
            "/api/shift-checklists",
            headers=self.auth_headers(worker_token),
            json={
                "phase": "end",
                "note": "Расход по двум машинам",
                "items": item_updates,
            },
        )
        self.assertEqual(end_response.status_code, 200, end_response.text)
        end_payload = end_response.json()
        self.assertEqual(end_payload["phase"], "end")
        self.assertTrue(all(item["startQty"] is not None for item in end_payload["items"]))

        worker_list = self.client.get("/api/shift-checklists", headers=self.auth_headers(worker_token))
        self.assertEqual(worker_list.status_code, 200, worker_list.text)
        self.assertEqual(len(worker_list.json()), 2)
        self.assertTrue(all(entry["workerId"] == "w1" for entry in worker_list.json()))

        owner_list = self.client.get("/api/shift-checklists", headers=self.auth_headers(owner_token))
        self.assertEqual(owner_list.status_code, 200, owner_list.text)
        self.assertEqual(len(owner_list.json()), 2)

        with SessionLocal() as db:
            notifications = db.scalars(
                select(Notification).where(Notification.recipient_role == "owner").order_by(Notification.created_at.desc())
            ).all()

        self.assertTrue(any("заполнил чек-лист начала смены" in item.message for item in notifications))
        self.assertTrue(any("заполнил чек-лист закрытия смены" in item.message for item in notifications))

    def test_admin_shift_inspection_sends_owner_photo_and_can_be_approved(self) -> None:
        from app.database import SessionLocal
        from app.models import Notification

        self.disable_owner_two_factor()
        self.set_primary_owner_telegram("123123123")
        owner_token = self.login_staff("owner", "owner")
        admin_token = self.login_staff("admin", "admin")

        created_items = []
        for payload in (
            {"name": "Активная пена", "qty": 20, "unit": "л", "unitPrice": 450, "category": "Химия"},
            {"name": "Микрофибра", "qty": 12, "unit": "шт", "unitPrice": 180, "category": "Расходники"},
        ):
            response = self.client.post("/api/stock-items", headers=self.auth_headers(owner_token), json=payload)
            self.assertEqual(response.status_code, 200, response.text)
            created_items.append(response.json())

        sent_photos: list[dict[str, object]] = []

        def fake_send_photo(chat_id: str | int, **kwargs) -> None:
            sent_photos.append({"chat_id": chat_id, **kwargs})

        with patch("app.main.send_telegram_photo", side_effect=fake_send_photo):
            create_response = self.client.post(
                "/api/admin/shift-inspections",
                headers=self.auth_headers(admin_token),
                json={
                    "floorPhotoUrl": "data:image/jpeg;base64,/9j/4AAQSkZJRgABAQAAAQABAAD/2w==",
                    "clothsReady": True,
                    "supplies": [{"stockItemId": item["id"], "checked": True} for item in created_items],
                    "masters": [{"workerId": "w1", "checked": True}],
                    "note": "Полы чистые, смена готова",
                },
            )
        self.assertEqual(create_response.status_code, 200, create_response.text)
        payload = create_response.json()
        self.assertEqual(payload["status"], "pending")
        self.assertTrue(any(photo["chat_id"] == "123123123" for photo in sent_photos))
        self.assertTrue(any("reply_markup" in photo for photo in sent_photos))

        review_response = self.client.post(
            f"/api/admin/shift-inspections/{payload['id']}/review",
            headers=self.auth_headers(owner_token),
            json={"action": "approved", "issueNote": ""},
        )
        self.assertEqual(review_response.status_code, 200, review_response.text)
        self.assertEqual(review_response.json()["status"], "approved")

        with SessionLocal() as db:
            notifications = db.scalars(
                select(Notification).where(Notification.recipient_role == "admin").order_by(Notification.created_at.desc())
            ).all()

        self.assertTrue(any("подтвердил открытие смены" in item.message for item in notifications))

    def test_admin_shift_inspection_list_uses_photo_endpoint(self) -> None:
        self.disable_owner_two_factor()
        owner_token = self.login_staff("owner", "owner")
        admin_token = self.login_staff("admin", "admin")

        response = self.client.post(
            "/api/stock-items",
            headers=self.auth_headers(owner_token),
            json={"name": "Микрофибра", "qty": 12, "unit": "шт", "unitPrice": 180, "category": "Расходники"},
        )
        self.assertEqual(response.status_code, 200, response.text)

        with patch("app.main.send_telegram_photo", side_effect=lambda *args, **kwargs: None):
            create_response = self.client.post(
                "/api/admin/shift-inspections",
                headers=self.auth_headers(admin_token),
                json={
                    "floorPhotoUrl": "data:image/jpeg;base64,/9j/4AAQSkZJRgABAQAAAQABAAD/2w==",
                    "clothsReady": True,
                    "supplies": [{"stockItemId": response.json()["id"], "checked": True}],
                    "masters": [{"workerId": "w1", "checked": True}],
                    "note": "Готово к старту",
                },
            )
        self.assertEqual(create_response.status_code, 200, create_response.text)

        admin_list = self.client.get("/api/admin/shift-inspections", headers=self.auth_headers(admin_token))
        self.assertEqual(admin_list.status_code, 200, admin_list.text)
        admin_payload = admin_list.json()
        self.assertEqual(len(admin_payload), 1)
        self.assertTrue(admin_payload[0]["floorPhotoUrl"].endswith(f"/api/admin/shift-inspections/{create_response.json()['id']}/photo"))

        owner_list = self.client.get("/api/admin/shift-inspections", headers=self.auth_headers(owner_token))
        self.assertEqual(owner_list.status_code, 200, owner_list.text)
        owner_payload = owner_list.json()
        self.assertEqual(len(owner_payload), 1)
        self.assertTrue(owner_payload[0]["floorPhotoUrl"].endswith(f"/api/admin/shift-inspections/{create_response.json()['id']}/photo"))

        photo_response = self.client.get(owner_payload[0]["floorPhotoUrl"], headers=self.auth_headers(owner_token))
        self.assertEqual(photo_response.status_code, 200, photo_response.text)
        self.assertEqual(photo_response.headers["content-type"], "image/jpeg")
        self.assertGreater(len(photo_response.content), 0)

    def test_bot_can_reject_admin_shift_with_issue_note(self) -> None:
        from bot import BotRuntime, process_telegram_update
        from app.database import SessionLocal
        from app.models import Notification

        self.disable_owner_two_factor()
        self.set_primary_owner_telegram("123123123")
        self.set_staff_telegram("admin", "456456456")
        owner_token = self.login_staff("owner", "owner")
        admin_token = self.login_staff("admin", "admin")

        response = self.client.post(
            "/api/stock-items",
            headers=self.auth_headers(owner_token),
            json={"name": "Микрофибра", "qty": 12, "unit": "шт", "unitPrice": 180, "category": "Расходники"},
        )
        self.assertEqual(response.status_code, 200, response.text)

        with patch("app.main.send_telegram_photo", side_effect=lambda *args, **kwargs: None):
            create_response = self.client.post(
                "/api/admin/shift-inspections",
                headers=self.auth_headers(admin_token),
                json={
                    "floorPhotoUrl": "data:image/jpeg;base64,/9j/4AAQSkZJRgABAQAAAQABAAD/2w==",
                    "clothsReady": True,
                    "supplies": [{"stockItemId": response.json()["id"], "checked": True}],
                    "masters": [{"workerId": "w1", "checked": True}],
                    "note": "Готово к старту",
                },
            )
        self.assertEqual(create_response.status_code, 200, create_response.text)
        inspection_id = create_response.json()["id"]

        telegram_calls: list[tuple[str, dict[str, object]]] = []

        def fake_telegram_call(_runtime, method: str, payload: dict[str, object] | None = None, **_kwargs):
            telegram_calls.append((method, payload or {}))
            return {}

        with patch("bot._build_runtime", return_value=BotRuntime(token="t", webapp_url="https://app.example", api_base="https://api.example")):
            with patch("bot._telegram_call", side_effect=fake_telegram_call):
                process_telegram_update(
                    {
                        "callback_query": {
                            "id": "cb1",
                            "data": f"shiftreject:{inspection_id}",
                            "message": {"chat": {"id": 123123123}},
                        }
                    }
                )
                process_telegram_update(
                    {
                        "message": {
                            "chat": {"id": 123123123},
                            "text": "Полы грязные у ворот",
                        }
                    }
                )

        with SessionLocal() as db:
            notifications = db.scalars(
                select(Notification).where(Notification.recipient_role == "admin").order_by(Notification.created_at.desc())
            ).all()

        self.assertTrue(any(method == "answerCallbackQuery" for method, _payload in telegram_calls))
        self.assertTrue(any("Полы грязные у ворот" in item.message for item in notifications))

    def test_admin_mark_read_all_affects_only_admin_notifications(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        owner_token = self.login_staff("owner", "owner") if False else None
        first_client_token, first_client_id = self.login_client(name="Alice", phone="+7 (999) 111-22-33")

        admin_notification = self.client.post(
            "/api/notifications",
            headers=self.auth_headers(admin_token),
            json={
                "recipientRole": "admin",
                "recipientId": None,
                "message": "admin-only",
                "read": False,
            },
        )
        self.assertEqual(admin_notification.status_code, 200, admin_notification.text)

        client_notification = self.client.post(
            "/api/notifications",
            headers=self.auth_headers(admin_token),
            json={
                "recipientRole": "client",
                "recipientId": first_client_id,
                "message": "client-only",
                "read": False,
            },
        )
        self.assertEqual(client_notification.status_code, 200, client_notification.text)

        read_all = self.client.post(
            "/api/notifications/read-all",
            headers=self.auth_headers(admin_token),
            json={"role": "admin"},
        )
        self.assertEqual(read_all.status_code, 200, read_all.text)

        admin_bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(admin_token))
        client_bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(first_client_token))
        self.assertEqual(admin_bootstrap.status_code, 200, admin_bootstrap.text)
        self.assertEqual(client_bootstrap.status_code, 200, client_bootstrap.text)
        self.assertTrue(all(item["read"] is True for item in admin_bootstrap.json()["notifications"]))
        self.assertTrue(any(item["read"] is False for item in client_bootstrap.json()["notifications"]))

    def test_admin_cannot_access_owner_only_endpoints(self) -> None:
        admin_token = self.login_staff("admin", "admin")

        create_worker = self.client.post(
            "/api/workers",
            headers=self.auth_headers(admin_token),
            json={
                "name": "????? ??????",
                "login": "newworker",
                "password": "password123",
                "percent": 30,
                "salaryBase": 0,
                "phone": "+7 (999) 333-44-55",
                "email": "worker@example.com",
                "telegramChatId": "",
            },
        )
        self.assertEqual(create_worker.status_code, 403, create_worker.text)

        stock_item = self.client.post(
            "/api/stock-items",
            headers=self.auth_headers(admin_token),
            json={
                "name": "???????",
                "qty": 10,
                "unit": "??",
                "unitPrice": 100,
                "category": "?????",
            },
        )
        self.assertEqual(stock_item.status_code, 403, stock_item.text)

        expense = self.client.post(
            "/api/expenses",
            headers=self.auth_headers(admin_token),
            json={
                "title": "??????",
                "amount": 10000,
                "category": "????",
                "date": self.next_active_date(),
                "note": "test",
            },
        )
        self.assertEqual(expense.status_code, 403, expense.text)


    def test_worker_can_update_only_own_assigned_booking_status_price_and_notes(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        worker_token = self.login_staff("ivan", "master")
        worker = self.get_staff(login="ivan")
        create_response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(admin_token),
            json={
                "clientId": "",
                "clientName": "Alice",
                "clientPhone": "+7 (999) 111-22-33",
                "service": "??????? ?????",
                "serviceId": "s1",
                "date": self.next_active_date(),
                "time": "18:00",
                "duration": 30,
                "price": 1200,
                "status": "scheduled",
                "workers": [{"workerId": worker["id"], "workerName": "????", "percent": 40}],
                "box": "???? 1",
                "paymentType": "cash",
                "car": "Lada Vesta",
                "plate": "A123BC",
            },
        )
        self.assertEqual(create_response.status_code, 200, create_response.text)
        booking_id = create_response.json()["id"]

        update_response = self.client.patch(
            f"/api/bookings/{booking_id}",
            headers=self.auth_headers(worker_token),
            json={"status": "in_progress", "price": 1500, "notes": "done prep"},
        )
        self.assertEqual(update_response.status_code, 200, update_response.text)
        payload = update_response.json()
        self.assertEqual(payload["status"], "in_progress")
        self.assertEqual(payload["price"], 1500)
        self.assertEqual(payload["notes"], "done prep")

    def test_worker_completion_creates_admin_notification_with_amount_client_and_service(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        worker_token = self.login_staff("ivan", "master")
        worker = self.get_staff(login="ivan")
        create_response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(admin_token),
            json={
                "clientId": "",
                "clientName": "Alice",
                "clientPhone": "+7 (999) 111-22-33",
                "service": "??????? ?????",
                "serviceId": "s1",
                "date": self.next_active_date(),
                "time": "18:30",
                "duration": 30,
                "price": 2100,
                "status": "scheduled",
                "workers": [{"workerId": worker["id"], "workerName": "????", "percent": 40}],
                "box": "???? 1",
                "paymentType": "cash",
                "car": "Lada Vesta",
                "plate": "A123BC",
            },
        )
        self.assertEqual(create_response.status_code, 200, create_response.text)
        booking_id = create_response.json()["id"]

        finish_response = self.client.patch(
            f"/api/bookings/{booking_id}",
            headers=self.auth_headers(worker_token),
            json={"status": "completed", "price": 2500, "paymentSettled": False},
        )
        self.assertEqual(finish_response.status_code, 200, finish_response.text)

        admin_bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(admin_token))
        self.assertEqual(admin_bootstrap.status_code, 200, admin_bootstrap.text)
        admin_messages = [item["message"] for item in admin_bootstrap.json()["notifications"] if item["recipientRole"] == "admin"]
        self.assertTrue(any("Alice" in message and "2500" in message and ("???????" in message or "???????" in message) for message in admin_messages))

    def test_worker_cannot_change_time_or_workers_even_on_own_booking(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        worker_token = self.login_staff("ivan", "master")
        worker = self.get_staff(login="ivan")
        create_response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(admin_token),
            json={
                "clientId": "",
                "clientName": "Alice",
                "clientPhone": "+7 (999) 111-22-33",
                "service": "??????? ?????",
                "serviceId": "s1",
                "date": self.next_active_date(),
                "time": "19:00",
                "duration": 30,
                "price": 1200,
                "status": "scheduled",
                "workers": [{"workerId": worker["id"], "workerName": "????", "percent": 40}],
                "box": "???? 1",
                "paymentType": "cash",
                "car": "Lada Vesta",
                "plate": "A123BC",
            },
        )
        self.assertEqual(create_response.status_code, 200, create_response.text)
        booking_id = create_response.json()["id"]

        change_time = self.client.patch(
            f"/api/bookings/{booking_id}",
            headers=self.auth_headers(worker_token),
            json={"time": "19:30"},
        )
        self.assertEqual(change_time.status_code, 403, change_time.text)

        change_workers = self.client.patch(
            f"/api/bookings/{booking_id}",
            headers=self.auth_headers(worker_token),
            json={"workers": []},
        )
        self.assertEqual(change_workers.status_code, 403, change_workers.text)

    def test_worker_must_specify_payment_state_when_completing_booking(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        worker_token = self.login_staff("ivan", "master")
        worker = self.get_staff(login="ivan")
        create_response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(admin_token),
            json={
                "clientId": "",
                "clientName": "Alice",
                "clientPhone": "+7 (999) 111-22-33",
                "service": "Комплексная мойка",
                "serviceId": "s1",
                "date": self.next_active_date(),
                "time": "19:00",
                "duration": 30,
                "price": 1800,
                "status": "scheduled",
                "workers": [{"workerId": worker["id"], "workerName": "Иван", "percent": 40}],
                "box": "Мойка самообслуживания",
                "paymentType": "cash",
                "paymentSettled": True,
                "car": "Lada Vesta",
                "plate": "A123BC",
            },
        )
        self.assertEqual(create_response.status_code, 200, create_response.text)
        booking_id = create_response.json()["id"]

        missing_status_response = self.client.patch(
            f"/api/bookings/{booking_id}",
            headers=self.auth_headers(worker_token),
            json={"status": "completed", "price": 1800},
        )
        self.assertEqual(missing_status_response.status_code, 400, missing_status_response.text)
        self.assertIn("оплатил ли клиент", missing_status_response.text.lower())

        missing_method_response = self.client.patch(
            f"/api/bookings/{booking_id}",
            headers=self.auth_headers(worker_token),
            json={"status": "completed", "price": 1800, "paymentSettled": True},
        )
        self.assertEqual(missing_method_response.status_code, 400, missing_method_response.text)
        self.assertIn("способ оплаты", missing_method_response.text.lower())

        unpaid_response = self.client.patch(
            f"/api/bookings/{booking_id}",
            headers=self.auth_headers(worker_token),
            json={"status": "completed", "price": 1800, "paymentSettled": False},
        )
        self.assertEqual(unpaid_response.status_code, 200, unpaid_response.text)
        self.assertEqual(unpaid_response.json()["paymentSettled"], False)

    def test_worker_can_save_only_own_profile(self) -> None:
        worker_token = self.login_staff("ivan", "master")
        worker = self.get_staff(login="ivan")
        other_worker = self.get_staff(login="oleg")

        own_response = self.client.put(
            f"/api/settings/workers/{worker['id']}/profile",
            headers=self.auth_headers(worker_token),
            json={
                "name": "???? ?????",
                "phone": "+7 (999) 444-55-66",
                "email": "ivan@example.com",
                "city": "??????",
                "experience": "5 ???",
                "specialty": "?????????",
                "about": "??????? ??????",
                "percent": 35,
            },
        )
        self.assertEqual(own_response.status_code, 200, own_response.text)
        self.assertEqual(own_response.json()["name"], "???? ?????")

        foreign_response = self.client.put(
            f"/api/settings/workers/{other_worker['id']}/profile",
            headers=self.auth_headers(worker_token),
            json={
                "name": "???? ?????",
                "phone": "+7 (999) 777-88-99",
                "email": "oleg@example.com",
                "city": "??????",
                "experience": "3 ????",
                "specialty": "?????",
                "about": "????? ???????",
                "percent": 20,
            },
        )
        self.assertEqual(foreign_response.status_code, 403, foreign_response.text)

    def test_worker_can_save_only_own_notification_settings(self) -> None:
        worker_token = self.login_staff("ivan", "master")
        worker = self.get_staff(login="ivan")
        other_worker = self.get_staff(login="oleg")

        own_response = self.client.put(
            f"/api/settings/workers/{worker['id']}/notifications",
            headers=self.auth_headers(worker_token),
            json={
                "newTask": True,
                "taskUpdate": False,
                "payment": True,
                "reminders": False,
                "sms": True,
            },
        )
        self.assertEqual(own_response.status_code, 200, own_response.text)
        self.assertEqual(own_response.json()["payment"], True)
        self.assertEqual(own_response.json()["taskUpdate"], False)

        foreign_response = self.client.put(
            f"/api/settings/workers/{other_worker['id']}/notifications",
            headers=self.auth_headers(worker_token),
            json={
                "newTask": False,
                "taskUpdate": False,
                "payment": False,
                "reminders": False,
                "sms": False,
            },
        )
        self.assertEqual(foreign_response.status_code, 403, foreign_response.text)

    def test_worker_mark_read_all_affects_only_own_notifications(self) -> None:
        self.disable_owner_two_factor()
        owner_token = self.login_staff("owner", "owner")
        worker_token = self.login_staff("ivan", "master")
        worker = self.get_staff(login="ivan")
        other_worker = self.get_staff(login="oleg")

        first_penalty = self.client.post(
            "/api/penalties",
            headers=self.auth_headers(owner_token),
            json={"workerId": worker["id"], "title": "?????????", "reason": "test"},
        )
        self.assertEqual(first_penalty.status_code, 200, first_penalty.text)
        second_penalty = self.client.post(
            "/api/penalties",
            headers=self.auth_headers(owner_token),
            json={"workerId": other_worker["id"], "title": "????????", "reason": "test"},
        )
        self.assertEqual(second_penalty.status_code, 200, second_penalty.text)

        before_other = self.count_worker_notifications(other_worker["id"])
        read_all = self.client.post(
            "/api/notifications/read-all",
            headers=self.auth_headers(worker_token),
            json={"role": "worker"},
        )
        self.assertEqual(read_all.status_code, 200, read_all.text)

        worker_bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(worker_token))
        self.assertEqual(worker_bootstrap.status_code, 200, worker_bootstrap.text)
        self.assertTrue(all(item["read"] is True for item in worker_bootstrap.json()["notifications"]))
        self.assertGreaterEqual(self.count_worker_notifications(other_worker["id"]), before_other)

    def test_worker_cannot_create_penalties(self) -> None:
        worker_token = self.login_staff("ivan", "master")
        other_worker = self.get_staff(login="oleg")
        response = self.client.post(
            "/api/penalties",
            headers=self.auth_headers(worker_token),
            json={"workerId": other_worker["id"], "title": "bad", "reason": "bad"},
        )
        self.assertEqual(response.status_code, 403, response.text)

    def test_worker_cannot_create_notifications_for_other_roles(self) -> None:
        worker_token = self.login_staff("ivan", "master")
        _, client_id = self.login_client(name="Alice", phone="+7 (999) 111-22-33")
        response = self.client.post(
            "/api/notifications",
            headers=self.auth_headers(worker_token),
            json={
                "recipientRole": "client",
                "recipientId": client_id,
                "message": "spoof",
                "read": False,
            },
        )
        self.assertEqual(response.status_code, 403, response.text)

    def test_worker_can_create_notification_for_assigned_client(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        worker_token = self.login_staff("ivan", "master")
        worker = self.get_staff(login="ivan")
        create_response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(admin_token),
            json={
                "clientId": "",
                "clientName": "Alice",
                "clientPhone": "+7 (999) 111-22-33",
                "service": "??????? ?????",
                "serviceId": "s1",
                "date": self.next_active_date(),
                "time": "20:00",
                "duration": 30,
                "price": 1200,
                "status": "scheduled",
                "workers": [{"workerId": worker["id"], "workerName": "????", "percent": 40}],
                "box": "???? 1",
                "paymentType": "cash",
                "car": "Lada Vesta",
                "plate": "A123BC",
            },
        )
        self.assertEqual(create_response.status_code, 200, create_response.text)
        client_id = create_response.json()["clientId"]

        response = self.client.post(
            "/api/notifications",
            headers=self.auth_headers(worker_token),
            json={
                "recipientRole": "client",
                "recipientId": client_id,
                "message": "??? ????? ????????.",
                "read": False,
            },
        )
        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["recipientId"], client_id)

    def test_worker_can_generate_telegram_link_code(self) -> None:
        worker_token = self.login_staff("ivan", "master")
        response = self.client.post(
            "/api/telegram/link-code",
            headers=self.auth_headers(worker_token),
        )
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(len(payload["code"]), 6)
        self.assertIn("expiresAt", payload)

    def test_telegram_webhook_rejects_invalid_secret(self) -> None:
        os.environ["WEBAPP_URL"] = "https://crm.example"
        os.environ["TELEGRAM_DELIVERY_MODE"] = "webhook"
        self.restart_app()

        response = self.client.post(
            "/api/telegram/webhook",
            headers={"X-Telegram-Bot-Api-Secret-Token": "invalid"},
            json={"update_id": 1, "message": {"chat": {"id": 123}, "text": "/start"}},
        )
        self.assertEqual(response.status_code, 401, response.text)

    def test_telegram_webhook_processes_update_with_valid_secret(self) -> None:
        os.environ["WEBAPP_URL"] = "https://crm.example"
        os.environ["TELEGRAM_DELIVERY_MODE"] = "webhook"
        self.restart_app()

        payload = {"update_id": 1, "message": {"chat": {"id": 123}, "text": "/start"}}
        with patch("app.main.process_telegram_update") as process_update:
            response = self.client.post(
                "/api/telegram/webhook",
                headers={"X-Telegram-Bot-Api-Secret-Token": self.telegram_webhook_secret()},
                json=payload,
            )
        self.assertEqual(response.status_code, 200, response.text)
        process_update.assert_called_once_with(payload)

    def test_client_bootstrap_contains_only_own_bookings_and_no_worker_directory(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        first_token, first_id = self.login_client(name="Alice", phone="+7 (999) 111-22-33")
        _second_token, second_id = self.login_client(name="Bob", phone="+7 (999) 222-33-44")
        booking_date = self.next_active_date()

        first_booking = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(admin_token),
            json={
                "clientId": first_id,
                "clientName": "Alice",
                "clientPhone": "+7 (999) 111-22-33",
                "service": "Base wash",
                "serviceId": "s1",
                "date": booking_date,
                "time": "10:00",
                "duration": 30,
                "price": 1200,
                "status": "scheduled",
                "workers": [],
                "box": "Р‘РѕРєСЃ 1",
                "paymentType": "cash",
                "car": "Lada Vesta",
                "plate": "A123BC",
            },
        )
        self.assertEqual(first_booking.status_code, 200, first_booking.text)

        second_booking = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(admin_token),
            json={
                "clientId": second_id,
                "clientName": "Bob",
                "clientPhone": "+7 (999) 222-33-44",
                "service": "Base wash",
                "serviceId": "s1",
                "date": booking_date,
                "time": "11:00",
                "duration": 30,
                "price": 1200,
                "status": "scheduled",
                "workers": [],
                "box": "Р‘РѕРєСЃ 1",
                "paymentType": "cash",
                "car": "Kia Rio",
                "plate": "B222BB",
            },
        )
        self.assertEqual(second_booking.status_code, 200, second_booking.text)

        bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(first_token))
        self.assertEqual(bootstrap.status_code, 200, bootstrap.text)
        payload = bootstrap.json()
        booking_ids = [item["id"] for item in payload["bookings"]]
        self.assertEqual(booking_ids, [first_booking.json()["id"]])
        self.assertEqual(payload["workers"], [])
        self.assertTrue(all(item["clientId"] == first_id for item in payload["bookings"]))

    def test_worker_bootstrap_contains_only_assigned_bookings(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        worker_token = self.login_staff("ivan", "master")
        first_worker = self.get_staff(login="ivan")
        second_worker = self.get_staff(login="oleg")
        booking_date = self.next_active_date()

        first_booking = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(admin_token),
            json={
                "clientId": "",
                "clientName": "Alice",
                "clientPhone": "+7 (999) 111-22-33",
                "service": "Base wash",
                "serviceId": "s1",
                "date": booking_date,
                "time": "12:00",
                "duration": 30,
                "price": 1200,
                "status": "scheduled",
                "workers": [{"workerId": first_worker["id"], "workerName": "Ivan", "percent": 40}],
                "box": "Р‘РѕРєСЃ 1",
                "paymentType": "cash",
                "car": "Lada Vesta",
                "plate": "A123BC",
            },
        )
        self.assertEqual(first_booking.status_code, 200, first_booking.text)

        second_booking = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(admin_token),
            json={
                "clientId": "",
                "clientName": "Bob",
                "clientPhone": "+7 (999) 222-33-44",
                "service": "Base wash",
                "serviceId": "s1",
                "date": booking_date,
                "time": "13:00",
                "duration": 30,
                "price": 1200,
                "status": "scheduled",
                "workers": [{"workerId": second_worker["id"], "workerName": "Oleg", "percent": 40}],
                "box": "Р‘РѕРєСЃ 1",
                "paymentType": "cash",
                "car": "Kia Rio",
                "plate": "B222BB",
            },
        )
        self.assertEqual(second_booking.status_code, 200, second_booking.text)

        bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(worker_token))
        self.assertEqual(bootstrap.status_code, 200, bootstrap.text)
        payload = bootstrap.json()
        booking_ids = [item["id"] for item in payload["bookings"]]
        self.assertEqual(booking_ids, [first_booking.json()["id"]])
        self.assertTrue(all(first_worker["id"] in [worker["workerId"] for worker in item["workers"]] for item in payload["bookings"]))

    def test_admin_can_update_booking_alias_fields_and_service_canonical_data(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        create_response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(admin_token),
            json={
                "clientId": "",
                "clientName": "Alice",
                "clientPhone": "+7 (999) 111-22-33",
                "service": "Base wash",
                "serviceId": "s1",
                "date": self.next_active_date(),
                "time": "14:00",
                "duration": 30,
                "price": 1200,
                "status": "scheduled",
                "workers": [],
                "box": "Р‘РѕРєСЃ 1",
                "paymentType": "cash",
                "car": "Lada Vesta",
                "plate": "A123BC",
            },
        )
        self.assertEqual(create_response.status_code, 200, create_response.text)
        booking_id = create_response.json()["id"]

        update_response = self.client.patch(
            f"/api/bookings/{booking_id}",
            headers=self.auth_headers(admin_token),
            json={
                "clientName": "Bob",
                "clientPhone": "8 (999) 444-55-66",
                "service": "Spoofed update",
                "serviceId": "s5",
                "paymentType": "card",
                "car": "Kia Rio",
                "plate": "B222BB",
            },
        )
        self.assertEqual(update_response.status_code, 200, update_response.text)
        payload = update_response.json()
        self.assertEqual(payload["clientName"], "Bob")
        self.assertEqual(payload["clientPhone"], "+7 (999) 444-55-66")
        self.assertEqual(payload["serviceId"], "s5")
        self.assertNotEqual(payload["service"], "Spoofed update")
        self.assertEqual(payload["duration"], 90)
        self.assertEqual(payload["price"], 4200)
        self.assertEqual(payload["paymentType"], "card")
        self.assertEqual(payload["car"], "Kia Rio")
        self.assertEqual(payload["plate"], "B222BB")

    def test_owner_stock_write_off_rejects_negative_qty(self) -> None:
        self.disable_owner_two_factor()
        owner_token = self.login_staff("owner", "owner")
        create_response = self.client.post(
            "/api/stock-items",
            headers=self.auth_headers(owner_token),
            json={
                "name": "Foam",
                "qty": 10,
                "unit": "pcs",
                "unitPrice": 100,
                "category": "chemistry",
            },
        )
        self.assertEqual(create_response.status_code, 200, create_response.text)
        item_id = create_response.json()["id"]

        write_off_response = self.client.post(
            f"/api/stock-items/{item_id}/write-off",
            headers=self.auth_headers(owner_token),
            json={"qty": -5},
        )
        self.assertEqual(write_off_response.status_code, 422, write_off_response.text)

        bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(owner_token))
        self.assertEqual(bootstrap.status_code, 200, bootstrap.text)
        item = next(stock for stock in bootstrap.json()["stockItems"] if stock["id"] == item_id)
        self.assertEqual(item["qty"], 10)

    def test_admin_can_read_targeted_admin_notifications(self) -> None:
        self.disable_owner_two_factor()
        owner_token = self.login_staff("owner", "owner")
        admin_token = self.login_staff("admin", "admin")
        admin_id = self.get_staff(login="admin")["id"]

        first_notification = self.client.post(
            "/api/notifications",
            headers=self.auth_headers(owner_token),
            json={
                "recipientRole": "admin",
                "recipientId": admin_id,
                "message": "Targeted admin message",
                "read": False,
            },
        )
        self.assertEqual(first_notification.status_code, 200, first_notification.text)

        second_notification = self.client.post(
            "/api/notifications",
            headers=self.auth_headers(owner_token),
            json={
                "recipientRole": "admin",
                "recipientId": admin_id,
                "message": "Another targeted admin message",
                "read": False,
            },
        )
        self.assertEqual(second_notification.status_code, 200, second_notification.text)

        mark_read = self.client.patch(
            f"/api/notifications/{first_notification.json()['id']}/read",
            headers=self.auth_headers(admin_token),
        )
        self.assertEqual(mark_read.status_code, 200, mark_read.text)

        mark_all = self.client.post(
            "/api/notifications/read-all",
            headers=self.auth_headers(admin_token),
            json={"role": "admin"},
        )
        self.assertEqual(mark_all.status_code, 200, mark_all.text)

        bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(admin_token))
        self.assertEqual(bootstrap.status_code, 200, bootstrap.text)
        notifications = {
            item["id"]: item
            for item in bootstrap.json()["notifications"]
            if item["recipientRole"] == "admin" and item["recipientId"] == admin_id
        }
        self.assertTrue(notifications[first_notification.json()["id"]]["read"])
        self.assertTrue(notifications[second_notification.json()["id"]]["read"])

    def test_deleting_client_removes_related_bookings_and_sessions(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        client_token, client_id = self.login_client(name="Alice", phone="+7 (999) 111-22-33")
        create_response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(admin_token),
            json={
                "clientId": client_id,
                "clientName": "Alice",
                "clientPhone": "+7 (999) 111-22-33",
                "service": "Base wash",
                "serviceId": "s1",
                "date": self.next_active_date(),
                "time": "15:00",
                "duration": 30,
                "price": 1200,
                "status": "scheduled",
                "workers": [],
                "box": "Р‘РѕРєСЃ 1",
                "paymentType": "cash",
                "car": "Lada Vesta",
                "plate": "A123BC",
            },
        )
        self.assertEqual(create_response.status_code, 200, create_response.text)
        booking_id = create_response.json()["id"]

        delete_response = self.client.delete(f"/api/clients/{client_id}", headers=self.auth_headers(admin_token))
        self.assertEqual(delete_response.status_code, 200, delete_response.text)
        self.assertEqual(self.count_client_sessions(client_id), 0)

        client_session = self.client.get("/api/auth/session", headers=self.auth_headers(client_token))
        self.assertEqual(client_session.status_code, 401, client_session.text)

        admin_bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(admin_token))
        self.assertEqual(admin_bootstrap.status_code, 200, admin_bootstrap.text)
        self.assertFalse(any(item["id"] == booking_id for item in admin_bootstrap.json()["bookings"]))

    def test_worker_cannot_message_client_from_only_completed_booking(self) -> None:
        admin_token = self.login_staff("admin", "admin")
        worker_token = self.login_staff("ivan", "master")
        worker = self.get_staff(login="ivan")
        create_response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(admin_token),
            json={
                "clientId": "",
                "clientName": "Alice",
                "clientPhone": "+7 (999) 111-22-33",
                "service": "Base wash",
                "serviceId": "s1",
                "date": self.next_active_date(),
                "time": "16:00",
                "duration": 30,
                "price": 1200,
                "status": "scheduled",
                "workers": [{"workerId": worker["id"], "workerName": "Ivan", "percent": 40}],
                "box": "Р‘РѕРєСЃ 1",
                "paymentType": "cash",
                "car": "Lada Vesta",
                "plate": "A123BC",
            },
        )
        self.assertEqual(create_response.status_code, 200, create_response.text)
        booking = create_response.json()

        complete_response = self.client.patch(
            f"/api/bookings/{booking['id']}",
            headers=self.auth_headers(admin_token),
            json={"status": "completed"},
        )
        self.assertEqual(complete_response.status_code, 200, complete_response.text)

        notification_response = self.client.post(
            "/api/notifications",
            headers=self.auth_headers(worker_token),
            json={
                "recipientRole": "client",
                "recipientId": booking["clientId"],
                "message": "Are you still coming?",
                "read": False,
            },
        )
        self.assertEqual(notification_response.status_code, 403, notification_response.text)

    def test_owner_database_reset_execute_requires_delay_after_approval(self) -> None:
        self.disable_owner_two_factor()
        self.set_primary_owner_telegram()
        owner_token = self.login_staff("owner", "owner")
        sent_messages: list[tuple[str, str]] = []

        def fake_send_message(chat_id: str, text: str) -> None:
            sent_messages.append((chat_id, text))

        with patch("app.main.send_telegram_message", side_effect=fake_send_message):
            start_response = self.client.post(
                "/api/owner/database-reset/start",
                headers=self.auth_headers(owner_token),
                json={"password": "owner"},
            )
            self.assertEqual(start_response.status_code, 200, start_response.text)

        self.assertEqual(len(sent_messages), 1)
        creator_code = self.extract_owner_reset_code(sent_messages[0][1])
        start_payload = start_response.json()
        approve_response = self.client.post(
            "/api/owner/database-reset/approve",
            headers=self.auth_headers(owner_token),
            json={
                "requestId": start_payload["requestId"],
                "creatorCode": creator_code,
                "confirmationPhrase": start_payload["confirmationPhrase"],
            },
        )
        self.assertEqual(approve_response.status_code, 200, approve_response.text)

        execute_response = self.client.post(
            "/api/owner/database-reset/execute",
            headers=self.auth_headers(owner_token),
            json={"requestId": start_payload["requestId"]},
        )
        self.assertEqual(execute_response.status_code, 409, execute_response.text)
        detail = execute_response.json()["detail"].lower()
        self.assertTrue("кнопка" in detail or "рєрѕрїрєр°" in detail, detail)

    def test_owner_database_reset_clears_operational_data_and_preserves_owners(self) -> None:
        from app.database import SessionLocal
        from app.models import (
            AppSetting,
            Booking,
            Box,
            Client,
            Expense,
            Notification,
            PayrollEntry,
            Service,
            StaffUser,
            StockItem,
        )

        self.disable_owner_two_factor()
        self.set_primary_owner_telegram()
        owner_token = self.login_staff("owner", "owner")
        admin_token = self.login_staff("admin", "admin")
        admin_bootstrap = self.client.get("/api/auth/session", headers=self.auth_headers(admin_token))
        self.assertEqual(admin_bootstrap.status_code, 200, admin_bootstrap.text)
        box_name = admin_bootstrap.json()["boxes"][0]["name"]
        client_token, client_id = self.login_client(name="Alice", phone="+7 (999) 111-22-33")
        client_session = self.client.get("/api/auth/session", headers=self.auth_headers(client_token))
        self.assertEqual(client_session.status_code, 200, client_session.text)

        booking_response = self.client.post(
            "/api/bookings",
            headers=self.auth_headers(admin_token),
            json={
                "clientId": client_id,
                "clientName": "Alice",
                "clientPhone": "+7 (999) 111-22-33",
                "service": "Base wash",
                "serviceId": "s1",
                "date": self.next_active_date(),
                "time": "12:00",
                "duration": 30,
                "price": 1200,
                "status": "scheduled",
                "workers": [],
                "box": box_name,
                "paymentType": "cash",
                "car": "Lada Vesta",
                "plate": "A123BC",
            },
        )
        self.assertEqual(booking_response.status_code, 200, booking_response.text)

        stock_response = self.client.post(
            "/api/stock-items",
            headers=self.auth_headers(owner_token),
            json={"name": "РЁР°РјРїСѓРЅСЊ", "qty": 5, "unit": "С€С‚", "unitPrice": 400, "category": "РҐРёРјРёСЏ"},
        )
        self.assertEqual(stock_response.status_code, 200, stock_response.text)

        expense_response = self.client.post(
            "/api/expenses",
            headers=self.auth_headers(owner_token),
            json={"title": "РџСЂРѕРІРµСЂРєР°", "amount": 900, "category": "РџСЂРѕС‡РµРµ", "date": self.next_active_date(), "note": ""},
        )
        self.assertEqual(expense_response.status_code, 200, expense_response.text)
        payroll_response = self.client.post(
            "/api/payroll/entries",
            headers=self.auth_headers(owner_token),
            json={"workerId": "w1", "kind": "bonus", "amount": 500, "note": "РџСЂРµРјРёСЏ"},
        )
        self.assertEqual(payroll_response.status_code, 200, payroll_response.text)

        sent_messages: list[tuple[str, str]] = []

        def fake_send_message(chat_id: str, text: str) -> None:
            sent_messages.append((chat_id, text))

        with patch("app.main.send_telegram_message", side_effect=fake_send_message):
            start_response = self.client.post(
                "/api/owner/database-reset/start",
                headers=self.auth_headers(owner_token),
                json={"password": "owner"},
            )
            self.assertEqual(start_response.status_code, 200, start_response.text)

        self.assertEqual(len(sent_messages), 1)
        start_payload = start_response.json()
        creator_code = self.extract_owner_reset_code(sent_messages[0][1])
        approve_response = self.client.post(
            "/api/owner/database-reset/approve",
            headers=self.auth_headers(owner_token),
            json={
                "requestId": start_payload["requestId"],
                "creatorCode": creator_code,
                "confirmationPhrase": start_payload["confirmationPhrase"],
            },
        )
        self.assertEqual(approve_response.status_code, 200, approve_response.text)

        self.force_owner_reset_ready()
        execute_response = self.client.post(
            "/api/owner/database-reset/execute",
            headers=self.auth_headers(owner_token),
            json={"requestId": start_payload["requestId"]},
        )
        self.assertEqual(execute_response.status_code, 200, execute_response.text)
        execute_payload = execute_response.json()
        self.assertGreaterEqual(execute_payload["preview"]["clientsDeleted"], 1)
        self.assertGreaterEqual(execute_payload["preview"]["bookingsDeleted"], 1)
        self.assertGreaterEqual(execute_payload["preview"]["stockItemsDeleted"], 1)
        self.assertGreaterEqual(execute_payload["preview"]["expensesDeleted"], 1)

        with SessionLocal() as db:
            owners = db.scalars(select(StaffUser).where(StaffUser.role == "owner")).all()
            self.assertEqual(len(owners), 2)
            self.assertEqual(len(db.scalars(select(StaffUser).where(StaffUser.role != "owner")).all()), 0)
            self.assertEqual(len(db.scalars(select(Client)).all()), 0)
            self.assertEqual(len(db.scalars(select(Booking)).all()), 0)
            self.assertEqual(len(db.scalars(select(StockItem)).all()), 0)
            self.assertEqual(len(db.scalars(select(Expense)).all()), 0)
            self.assertEqual(len(db.scalars(select(PayrollEntry)).all()), 0)
            self.assertEqual(len(db.scalars(select(Notification)).all()), 0)
            self.assertGreater(len(db.scalars(select(Service)).all()), 0)
            self.assertGreater(len(db.scalars(select(Box)).all()), 0)
            self.assertIsNone(db.get(AppSetting, "owner_database_reset"))

        owner_session = self.client.get("/api/auth/session", headers=self.auth_headers(owner_token))
        self.assertEqual(owner_session.status_code, 200, owner_session.text)

        admin_relogin = self.client.post("/api/auth/staff/login", json={"login": "admin", "password": "admin"})
        self.assertEqual(admin_relogin.status_code, 401, admin_relogin.text)

    def test_normalize_service_and_box_resources_handles_legacy_null_box_fields(self) -> None:
        from app.main import DETAILING_BOX_NAME, WASH_BOX_NAMES, _normalize_service_and_box_resources
        from app.models import Box, Service

        class FakeScalarResult:
            def __init__(self, items: list[object]) -> None:
                self._items = items

            def all(self) -> list[object]:
                return self._items

        class FakeSession:
            def __init__(self, services: list[Service], boxes: list[Box]) -> None:
                self.services = services
                self.boxes = boxes
                self.flushed = False

            def scalars(self, statement):
                entity = statement.column_descriptions[0]["entity"]
                if entity is Service:
                    return FakeScalarResult(self.services)
                if entity is Box:
                    return FakeScalarResult(self.boxes)
                raise AssertionError(f"Unexpected entity: {entity}")

            def add(self, _item: object) -> None:
                return None

            def flush(self) -> None:
                self.flushed = True

        legacy_boxes = [
            Box(
                id="legacy-wash",
                name=None,
                resource_group="wash",
                price_per_hour=500,
                active=True,
                description=None,
            ),
            Box(
                id="legacy-detail",
                name=None,
                resource_group="detailing",
                price_per_hour=700,
                active=True,
                description=None,
            ),
        ]
        fake_db = FakeSession([], legacy_boxes)

        _normalize_service_and_box_resources(fake_db)

        self.assertTrue(fake_db.flushed)
        self.assertEqual(legacy_boxes[0].name, WASH_BOX_NAMES[0])
        self.assertEqual(legacy_boxes[1].name, DETAILING_BOX_NAME)
        self.assertEqual(legacy_boxes[1].description, "Отдельное помещение для детейлинга")
        self.assertGreaterEqual(len(fake_db.boxes), 3)


if __name__ == "__main__":
    unittest.main()


