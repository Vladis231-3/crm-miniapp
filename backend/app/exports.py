from __future__ import annotations

import io
import os
from dataclasses import dataclass
from datetime import datetime, timedelta
from functools import lru_cache
from pathlib import Path
from typing import Any, Literal
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, Spacer, TableStyle

from .complaints import adjusted_booking_percent, complaint_status_for_percent
from .models import Booking, Expense, Penalty, Service, StaffUser, StockItem


ExportKind = Literal["report", "pdf"]
STATUS_LABELS = {
    "scheduled": "Запланировано",
    "in_progress": "В работе",
    "completed": "Завершено",
    "cancelled": "Отменено",
    "admin_review": "На уточнении админа",
}
PAYMENT_LABELS = {
    "cash": "Наличные",
    "card": "Карта",
    "online": "Онлайн",
}


@dataclass(frozen=True)
class ExportMetric:
    label: str
    value: str


@dataclass(frozen=True)
class OwnerExportData:
    owner_name: str
    company_name: str
    generated_at: datetime
    period_from: str
    period_to: str
    metrics: list[ExportMetric]
    service_rows: list[list[Any]]
    payroll_rows: list[list[Any]]
    client_rows: list[list[Any]]
    booking_rows: list[list[Any]]
    expense_category_rows: list[list[Any]]
    expense_rows: list[list[Any]]
    stock_rows: list[list[Any]]
    complaint_rows: list[list[Any]]


@dataclass(frozen=True)
class GeneratedExport:
    file_name: str
    media_type: str
    content: bytes
    telegram_caption: str


ReportPeriod = Literal["daily", "weekly"]
ReportSegment = Literal["wash", "detailing"]


@dataclass(frozen=True)
class OwnerSummaryReport:
    title: str
    message: str


@dataclass(frozen=True)
class OwnerSummaryContext:
    company_name: str
    generated_at: datetime
    period: ReportPeriod
    segment: ReportSegment
    period_label: str
    segment_label: str
    title: str
    filtered: list[tuple[Booking, Service | None]]


@dataclass(frozen=True)
class OwnerSummaryExportData:
    owner_name: str
    company_name: str
    title: str
    generated_at: datetime
    period_label: str
    segment_label: str
    metrics: list[ExportMetric]
    status_rows: list[list[Any]]
    payment_rows: list[list[Any]]
    day_rows: list[list[Any]]
    box_rows: list[list[Any]]
    service_rows: list[list[Any]]
    client_rows: list[list[Any]]
    worker_rows: list[list[Any]]
    booking_rows: list[list[Any]]
    action_rows: list[list[Any]]


def build_owner_summary_report(
    *,
    company_name: str,
    bookings: list[Booking],
    services: list[Service],
    period: ReportPeriod,
    segment: ReportSegment,
    now: datetime | None = None,
) -> OwnerSummaryReport:
    context = _build_owner_summary_context(
        company_name=company_name,
        bookings=bookings,
        services=services,
        period=period,
        segment=segment,
        now=now,
    )
    filtered = context.filtered
    header = _summary_header(context)
    if not filtered:
        return OwnerSummaryReport(title=context.title, message=f"{header}\nЗаписей за выбранный период нет.")

    completed = [booking for booking, _service in filtered if booking.status == "completed"]
    scheduled = [booking for booking, _service in filtered if booking.status == "scheduled"]
    in_progress = [booking for booking, _service in filtered if booking.status == "in_progress"]
    admin_review = [booking for booking, _service in filtered if booking.status == "admin_review"]
    cancelled = [booking for booking, _service in filtered if booking.status == "cancelled"]
    revenue = sum(booking.price for booking in completed)
    avg_check = round(revenue / len(completed)) if completed else 0

    service_rollup: dict[str, dict[str, int | str]] = {}
    for booking, service in filtered:
        service_name = service.name if service is not None else booking.service
        row = service_rollup.setdefault(service_name, {"name": service_name, "total": 0, "completed": 0, "revenue": 0})
        row["total"] = int(row["total"]) + 1
        if booking.status == "completed":
            row["completed"] = int(row["completed"]) + 1
            row["revenue"] = int(row["revenue"]) + booking.price

    lines = [
        header,
        f"Всего записей: {len(filtered)}",
        f"Завершено: {len(completed)}",
        f"Активно: {len(scheduled) + len(in_progress)}",
        f"На уточнении админа: {len(admin_review)}",
        f"Отменено: {len(cancelled)}",
        f"Выручка: {_format_money(revenue)}",
        f"Средний чек: {_format_money(avg_check)}",
    ]

    top_services = sorted(
        service_rollup.values(),
        key=lambda item: (int(item["revenue"]), int(item["completed"]), int(item["total"]), str(item["name"])),
        reverse=True,
    )
    if top_services:
        lines.append("Топ услуг:")
        for index, row in enumerate(top_services[:3], start=1):
            lines.append(
                f"{index}. {row['name']} - {row['completed']} заверш., {_format_money(int(row['revenue']))}"
            )

    if admin_review:
        lines.append("Требуют закрытия администратором:")
        for booking in sorted(admin_review, key=_booking_sort_key, reverse=True)[:3]:
            lines.append(f"- {booking.client_name} - {booking.date} {booking.time} - {booking.service}")

    return OwnerSummaryReport(title=context.title, message="\n".join(lines))


def build_owner_summary_export(
    *,
    owner: StaffUser,
    company_name: str,
    bookings: list[Booking],
    services: list[Service],
    period: ReportPeriod,
    segment: ReportSegment,
    now: datetime | None = None,
) -> GeneratedExport:
    context = _build_owner_summary_context(
        company_name=company_name,
        bookings=bookings,
        services=services,
        period=period,
        segment=segment,
        now=now,
    )
    data = _build_owner_summary_export_data(owner_name=owner.name.strip() or owner.login, context=context)
    slug = data.generated_at.strftime("%Y-%m-%d-%H%M")
    file_name = f"owner-summary-{period}-{segment}-{slug}.xlsx"
    metrics = {metric.label: metric.value for metric in data.metrics}
    caption = (
        f"{data.company_name}\n"
        f"{data.title}\n"
        f"Период: {data.period_label}\n"
        f"Выручка: {metrics.get('Выручка завершённых', '0 руб.')}\n"
        f"Всего записей: {metrics.get('Всего записей', '0')}\n"
        "Подробный Excel-файл во вложении."
    )
    return GeneratedExport(
        file_name=file_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content=_render_owner_summary_excel_report(data),
        telegram_caption=caption,
    )


def _build_owner_summary_context(
    *,
    company_name: str,
    bookings: list[Booking],
    services: list[Service],
    period: ReportPeriod,
    segment: ReportSegment,
    now: datetime | None = None,
) -> OwnerSummaryContext:
    generated_at = now or datetime.now().astimezone()
    if generated_at.tzinfo is None:
        generated_at = generated_at.astimezone()
    period_start, period_end, period_title = _summary_period_bounds(period, generated_at)
    period_label = _summary_period_label(period_start, period_end)
    segment_label = "Мойка" if segment == "wash" else "Детейлинг"
    service_map = {service.id: service for service in services}

    filtered: list[tuple[Booking, Service | None]] = []
    for booking in bookings:
        booking_dt = _booking_datetime(booking)
        if booking_dt is None:
            continue
        booking_local = _as_local_datetime(booking_dt, generated_at)
        if booking_local < period_start or booking_local >= period_end:
            continue
        service = service_map.get(booking.service_id)
        if not _booking_matches_segment(booking, service, segment):
            continue
        filtered.append((booking, service))

    return OwnerSummaryContext(
        company_name=company_name.strip() or "ATMOSFERA",
        generated_at=generated_at,
        period=period,
        segment=segment,
        period_label=period_label,
        segment_label=segment_label,
        title=f"{period_title} отчёт по направлению: {segment_label}",
        filtered=filtered,
    )


def _summary_header(context: OwnerSummaryContext) -> str:
    return f"{context.company_name}\n{context.title}\nПериод: {context.period_label}"


def _build_owner_summary_export_data(*, owner_name: str, context: OwnerSummaryContext) -> OwnerSummaryExportData:
    sorted_items = sorted(context.filtered, key=lambda item: _booking_sort_key(item[0]), reverse=True)
    bookings = [booking for booking, _service in sorted_items]
    completed = [booking for booking in bookings if booking.status == "completed"]
    scheduled = [booking for booking in bookings if booking.status == "scheduled"]
    in_progress = [booking for booking in bookings if booking.status == "in_progress"]
    admin_review = [booking for booking in bookings if booking.status == "admin_review"]
    cancelled = [booking for booking in bookings if booking.status == "cancelled"]
    active = [booking for booking in bookings if booking.status in {"scheduled", "in_progress"}]
    open_value = sum(booking.price for booking in bookings if booking.status in {"scheduled", "in_progress", "admin_review"})
    revenue = sum(booking.price for booking in completed)
    avg_check = round(revenue / len(completed)) if completed else 0
    total_minutes = sum(booking.duration for booking in bookings)
    unique_clients = {
        booking.client_id or f"{booking.client_name}|{booking.client_phone}"
        for booking in bookings
    }
    unique_workers = {
        link.worker_id or link.worker_name
        for booking in bookings
        for link in booking.worker_links
        if (link.worker_id or link.worker_name)
    }
    unique_boxes = {booking.box.strip() for booking in bookings if booking.box.strip()}

    metrics = [
        ExportMetric("Всего записей", str(len(bookings))),
        ExportMetric("Завершено", str(len(completed))),
        ExportMetric("Запланировано", str(len(scheduled))),
        ExportMetric("В работе", str(len(in_progress))),
        ExportMetric("На уточнении админа", str(len(admin_review))),
        ExportMetric("Отменено", str(len(cancelled))),
        ExportMetric("Выручка завершённых", _format_money(revenue)),
        ExportMetric("Потенциал открытых записей", _format_money(open_value)),
        ExportMetric("Средний чек", _format_money(avg_check)),
        ExportMetric("Уникальных клиентов", str(len(unique_clients))),
        ExportMetric("Задействовано сотрудников", str(len(unique_workers))),
        ExportMetric("Задействовано боксов", str(len(unique_boxes))),
        ExportMetric("Общая загрузка, мин", str(total_minutes)),
        ExportMetric("Общая загрузка, ч", f"{round(total_minutes / 60, 1):.1f}"),
    ]

    status_rows: list[list[Any]] = []
    for status_key in ["completed", "scheduled", "in_progress", "admin_review", "cancelled"]:
        status_bookings = [booking for booking in bookings if booking.status == status_key]
        status_rows.append([
            STATUS_LABELS.get(status_key, status_key),
            len(status_bookings),
            sum(booking.price for booking in status_bookings),
            sum(booking.duration for booking in status_bookings),
            round((len(status_bookings) / len(bookings)) * 100) if bookings else 0,
        ])

    payment_rollup: dict[str, dict[str, int | str]] = {}
    for booking in bookings:
        payment_label = PAYMENT_LABELS.get(booking.payment_type, booking.payment_type or "Не указано")
        row = payment_rollup.setdefault(
            payment_label,
            {"label": payment_label, "total": 0, "completed": 0, "revenue": 0, "open_amount": 0},
        )
        row["total"] = int(row["total"]) + 1
        if booking.status == "completed":
            row["completed"] = int(row["completed"]) + 1
            row["revenue"] = int(row["revenue"]) + booking.price
        elif booking.status != "cancelled":
            row["open_amount"] = int(row["open_amount"]) + booking.price
    payment_rows = [
        [row["label"], row["total"], row["completed"], row["revenue"], row["open_amount"]]
        for row in payment_rollup.values()
    ]
    payment_rows.sort(key=lambda item: (item[3], item[1], item[0]), reverse=True)

    day_rollup: dict[str, dict[str, Any]] = {}
    box_rollup: dict[str, dict[str, Any]] = {}
    service_rollup: dict[str, dict[str, Any]] = {}
    client_rollup: dict[str, dict[str, Any]] = {}
    worker_rollup: dict[str, dict[str, Any]] = {}
    action_rows: list[list[Any]] = []
    booking_rows: list[list[Any]] = []

    for booking, service in sorted_items:
        booking_dt = _booking_datetime(booking)
        booking_key = booking.date
        booking_vehicle = " / ".join(part for part in [booking.car or "", booking.plate or ""] if part)
        service_category = (service.category if service is not None else "").strip() or "Без категории"
        worker_names = ", ".join(f"{link.worker_name} ({link.percent}%)" for link in booking.worker_links) or "Не назначены"

        day_row = day_rollup.setdefault(
            booking_key,
            {
                "date": booking.date,
                "sort_key": booking_dt or booking.created_at,
                "total": 0,
                "completed": 0,
                "active": 0,
                "admin_review": 0,
                "cancelled": 0,
                "revenue": 0,
                "minutes": 0,
                "clients": set(),
                "boxes": set(),
            },
        )
        day_row["total"] += 1
        day_row["minutes"] += booking.duration
        day_row["clients"].add(booking.client_id or f"{booking.client_name}|{booking.client_phone}")
        if booking.box.strip():
            day_row["boxes"].add(booking.box.strip())
        if booking.status == "completed":
            day_row["completed"] += 1
            day_row["revenue"] += booking.price
        elif booking.status in {"scheduled", "in_progress"}:
            day_row["active"] += 1
        elif booking.status == "admin_review":
            day_row["admin_review"] += 1
        elif booking.status == "cancelled":
            day_row["cancelled"] += 1

        box_row = box_rollup.setdefault(
            booking.box or "Не указан",
            {
                "name": booking.box or "Не указан",
                "total": 0,
                "completed": 0,
                "active": 0,
                "admin_review": 0,
                "cancelled": 0,
                "revenue": 0,
                "minutes": 0,
                "last": None,
            },
        )
        box_row["total"] += 1
        box_row["minutes"] += booking.duration
        if booking.status == "completed":
            box_row["completed"] += 1
            box_row["revenue"] += booking.price
        elif booking.status in {"scheduled", "in_progress"}:
            box_row["active"] += 1
        elif booking.status == "admin_review":
            box_row["admin_review"] += 1
        elif booking.status == "cancelled":
            box_row["cancelled"] += 1
        if booking_dt is not None and (box_row["last"] is None or booking_dt > box_row["last"]):
            box_row["last"] = booking_dt

        service_key = booking.service_id or booking.service
        service_row = service_rollup.setdefault(
            service_key,
            {
                "category": service_category,
                "name": service.name if service is not None else booking.service,
                "total": 0,
                "completed": 0,
                "active": 0,
                "admin_review": 0,
                "cancelled": 0,
                "revenue": 0,
                "last": None,
            },
        )
        service_row["total"] += 1
        if booking.status == "completed":
            service_row["completed"] += 1
            service_row["revenue"] += booking.price
        elif booking.status in {"scheduled", "in_progress"}:
            service_row["active"] += 1
        elif booking.status == "admin_review":
            service_row["admin_review"] += 1
        elif booking.status == "cancelled":
            service_row["cancelled"] += 1
        if booking_dt is not None and (service_row["last"] is None or booking_dt > service_row["last"]):
            service_row["last"] = booking_dt

        client_key = booking.client_id or f"{booking.client_name}|{booking.client_phone}"
        client_row = client_rollup.setdefault(
            client_key,
            {
                "name": booking.client_name,
                "phone": booking.client_phone,
                "vehicle": booking_vehicle,
                "total": 0,
                "completed": 0,
                "active": 0,
                "admin_review": 0,
                "cancelled": 0,
                "revenue": 0,
                "first": None,
                "last": None,
            },
        )
        client_row["total"] += 1
        if booking.status == "completed":
            client_row["completed"] += 1
            client_row["revenue"] += booking.price
        elif booking.status in {"scheduled", "in_progress"}:
            client_row["active"] += 1
        elif booking.status == "admin_review":
            client_row["admin_review"] += 1
        elif booking.status == "cancelled":
            client_row["cancelled"] += 1
        if booking_vehicle and not client_row["vehicle"]:
            client_row["vehicle"] = booking_vehicle
        if booking_dt is not None and (client_row["first"] is None or booking_dt < client_row["first"]):
            client_row["first"] = booking_dt
        if booking_dt is not None and (client_row["last"] is None or booking_dt > client_row["last"]):
            client_row["last"] = booking_dt

        for link in booking.worker_links:
            worker_key = link.worker_id or link.worker_name
            worker_row = worker_rollup.setdefault(
                worker_key,
                {
                    "name": link.worker_name or link.worker_id or "Не назначен",
                    "total": 0,
                    "completed": 0,
                    "active": 0,
                    "admin_review": 0,
                    "cancelled": 0,
                    "revenue": 0,
                    "earned": 0,
                    "minutes": 0,
                    "percent_total": 0,
                    "percent_count": 0,
                    "last": None,
                },
            )
            worker_row["total"] += 1
            worker_row["minutes"] += booking.duration
            worker_row["percent_total"] += link.percent
            worker_row["percent_count"] += 1
            if booking.status == "completed":
                worker_row["completed"] += 1
                worker_row["revenue"] += booking.price
                worker_row["earned"] += round(booking.price * link.percent / 100)
            elif booking.status in {"scheduled", "in_progress"}:
                worker_row["active"] += 1
            elif booking.status == "admin_review":
                worker_row["admin_review"] += 1
            elif booking.status == "cancelled":
                worker_row["cancelled"] += 1
            if booking_dt is not None and (worker_row["last"] is None or booking_dt > worker_row["last"]):
                worker_row["last"] = booking_dt

        if booking.status == "admin_review":
            action_rows.append([
                "Закрыть запись у администратора",
                STATUS_LABELS.get(booking.status, booking.status),
                booking.client_name,
                booking.service,
                f"{booking.date} {booking.time}",
                booking.box,
                worker_names,
                (booking.notes or "").strip(),
            ])
        elif booking.status in {"scheduled", "in_progress"} and not booking.worker_links:
            action_rows.append([
                "Назначить мастера",
                STATUS_LABELS.get(booking.status, booking.status),
                booking.client_name,
                booking.service,
                f"{booking.date} {booking.time}",
                booking.box,
                worker_names,
                (booking.notes or "").strip(),
            ])

        booking_rows.append([
            _format_datetime(booking.created_at),
            f"{booking.date} {booking.time}",
            STATUS_LABELS.get(booking.status, booking.status),
            booking.client_name,
            booking.client_phone,
            booking_vehicle,
            booking.service,
            service_category,
            booking.box,
            PAYMENT_LABELS.get(booking.payment_type, booking.payment_type),
            booking.duration,
            booking.price,
            worker_names,
            (booking.notes or "").strip(),
        ])

    day_rows = [
        [
            row["date"],
            row["total"],
            row["completed"],
            row["active"],
            row["admin_review"],
            row["cancelled"],
            row["revenue"],
            round(row["revenue"] / row["completed"]) if row["completed"] else 0,
            row["minutes"],
            len(row["clients"]),
            len(row["boxes"]),
        ]
        for row in day_rollup.values()
    ]
    day_rows.sort(key=lambda item: _parse_date_for_sort(item[0]))

    box_rows = [
        [
            row["name"],
            row["total"],
            row["completed"],
            row["active"],
            row["admin_review"],
            row["cancelled"],
            row["revenue"],
            round(row["revenue"] / row["completed"]) if row["completed"] else 0,
            row["minutes"],
            _format_datetime(row["last"]),
        ]
        for row in box_rollup.values()
    ]
    box_rows.sort(key=lambda item: (item[6], item[2], item[1], item[0]), reverse=True)

    service_rows = [
        [
            row["category"],
            row["name"],
            row["total"],
            row["completed"],
            row["active"],
            row["admin_review"],
            row["cancelled"],
            row["revenue"],
            round(row["revenue"] / row["completed"]) if row["completed"] else 0,
            _format_datetime(row["last"]),
        ]
        for row in service_rollup.values()
    ]
    service_rows.sort(key=lambda item: (item[7], item[3], item[2], item[1]), reverse=True)

    client_rows = [
        [
            row["name"],
            row["phone"],
            row["vehicle"],
            row["total"],
            row["completed"],
            row["active"],
            row["admin_review"],
            row["cancelled"],
            row["revenue"],
            _format_datetime(row["first"]),
            _format_datetime(row["last"]),
        ]
        for row in client_rollup.values()
    ]
    client_rows.sort(key=lambda item: (item[8], item[4], item[3], item[0]), reverse=True)

    worker_rows = [
        [
            row["name"],
            row["total"],
            row["completed"],
            row["active"],
            row["admin_review"],
            row["cancelled"],
            row["revenue"],
            row["earned"],
            round(row["percent_total"] / row["percent_count"]) if row["percent_count"] else 0,
            row["minutes"],
            _format_datetime(row["last"]),
        ]
        for row in worker_rollup.values()
    ]
    worker_rows.sort(key=lambda item: (item[7], item[6], item[1], item[0]), reverse=True)

    return OwnerSummaryExportData(
        owner_name=owner_name,
        company_name=context.company_name,
        title=context.title,
        generated_at=context.generated_at,
        period_label=context.period_label,
        segment_label=context.segment_label,
        metrics=metrics,
        status_rows=status_rows,
        payment_rows=payment_rows,
        day_rows=day_rows,
        box_rows=box_rows,
        service_rows=service_rows,
        client_rows=client_rows,
        worker_rows=worker_rows,
        booking_rows=booking_rows,
        action_rows=action_rows,
    )


def _summary_period_bounds(period: ReportPeriod, current: datetime) -> tuple[datetime, datetime, str]:
    end_at = current.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)
    if period == "daily":
        return end_at - timedelta(days=1), end_at, "Ежедневный"
    return end_at - timedelta(days=7), end_at, "Еженедельный"


def _summary_period_label(period_start: datetime, period_end: datetime) -> str:
    last_day = period_end - timedelta(days=1)
    if period_start.date() == last_day.date():
        return period_start.strftime("%d.%m.%Y")
    return f"{period_start.strftime('%d.%m.%Y')} - {last_day.strftime('%d.%m.%Y')}"


def _booking_matches_segment(booking: Booking, service: Service | None, segment: ReportSegment) -> bool:
    if service is not None and service.category:
        category = service.category.strip().lower()
        if segment == "wash":
            return "мойк" in category
        return "детейл" in category

    fallback = booking.service.strip().lower()
    if segment == "wash":
        return "мойк" in fallback
    return "детейл" in fallback


def build_owner_export(
    *,
    kind: ExportKind,
    owner: StaffUser,
    company_name: str,
    bookings: list[Booking],
    expenses: list[Expense],
    penalties: list[Penalty],
    workers: list[StaffUser],
    stock_items: list[StockItem],
    services: list[Service],
) -> GeneratedExport:
    data = _build_export_data(
        owner=owner,
        company_name=company_name,
        bookings=bookings,
        expenses=expenses,
        penalties=penalties,
        workers=workers,
        stock_items=stock_items,
        services=services,
    )
    slug = data.generated_at.strftime("%Y-%m-%d-%H%M")
    if kind == "report":
        return GeneratedExport(
            file_name=f"owner-report-{slug}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content=_render_excel_report(data),
            telegram_caption=f"Подробный Excel-отчёт владельца {data.company_name}",
        )
    return GeneratedExport(
        file_name=f"owner-report-{slug}.pdf",
        media_type="application/pdf",
        content=_render_pdf_report(data),
        telegram_caption=f"Подробный PDF-отчёт владельца {data.company_name}",
    )


def _build_export_data(
    *,
    owner: StaffUser,
    company_name: str,
    bookings: list[Booking],
    expenses: list[Expense],
    penalties: list[Penalty],
    workers: list[StaffUser],
    stock_items: list[StockItem],
    services: list[Service],
) -> OwnerExportData:
    generated_at = datetime.now().astimezone()
    bookings = sorted(bookings, key=_booking_sort_key, reverse=True)
    completed = [item for item in bookings if item.status == "completed"]
    in_progress = [item for item in bookings if item.status == "in_progress"]
    scheduled = [item for item in bookings if item.status == "scheduled"]
    cancelled = [item for item in bookings if item.status == "cancelled"]

    revenue = sum(item.price for item in completed)
    total_expenses = sum(item.amount for item in expenses)
    profit = revenue - total_expenses
    margin = round((profit / revenue) * 100) if revenue else 0
    avg_check = round(revenue / len(completed)) if completed else 0
    completion_rate = round((len(completed) / len(bookings)) * 100) if bookings else 0
    cancellation_rate = round((len(cancelled) / len(bookings)) * 100) if bookings else 0
    stock_value = sum(item.qty * item.unit_price for item in stock_items)

    booking_dates = [_booking_datetime(item) for item in bookings if _booking_datetime(item) is not None]
    period_from = min(booking_dates).strftime("%d.%m.%Y %H:%M") if booking_dates else "Нет данных"
    period_to = max(booking_dates).strftime("%d.%m.%Y %H:%M") if booking_dates else "Нет данных"

    service_map = {service.id: service for service in services}
    service_rollup: dict[str, dict[str, Any]] = {}
    for booking in bookings:
        service = service_map.get(booking.service_id)
        key = booking.service_id or booking.service
        row = service_rollup.setdefault(
            key,
            {
                "category": service.category if service else "Без категории",
                "name": service.name if service else booking.service,
                "total": 0,
                "completed": 0,
                "active": 0,
                "cancelled": 0,
                "revenue": 0,
                "last": None,
            },
        )
        row["total"] += 1
        if booking.status == "completed":
            row["completed"] += 1
            row["revenue"] += booking.price
        elif booking.status in {"scheduled", "in_progress"}:
            row["active"] += 1
        elif booking.status == "cancelled":
            row["cancelled"] += 1
        dt = _booking_datetime(booking)
        if dt and (row["last"] is None or dt > row["last"]):
            row["last"] = dt
    service_rows = [
        [
            row["category"], row["name"], row["total"], row["completed"], row["active"], row["cancelled"],
            row["revenue"], round(row["revenue"] / row["completed"]) if row["completed"] else 0,
            _format_datetime(row["last"]),
        ]
        for row in service_rollup.values()
    ]
    service_rows.sort(key=lambda item: (item[6], item[3], item[2], item[1]), reverse=True)

    complaints_by_worker: dict[str, list[Penalty]] = {worker.id: [] for worker in workers}
    for penalty in penalties:
        complaints_by_worker.setdefault(penalty.worker_id, []).append(penalty)

    payroll_rows: list[list[Any]] = []
    total_payroll = 0
    active_assignments = scheduled + in_progress
    for worker in workers:
        worker_completed = [booking for booking in completed if any(link.worker_id == worker.id for link in booking.worker_links)]
        worker_active = [booking for booking in active_assignments if any(link.worker_id == worker.id for link in booking.worker_links)]
        worker_penalties = complaints_by_worker.get(worker.id, [])
        earned = 0
        for booking in worker_completed:
            link = next((item for item in booking.worker_links if item.worker_id == worker.id), None)
            percent = adjusted_booking_percent(
                link.percent if link is not None else 0,
                worker_penalties,
                date_value=booking.date,
                time_value=booking.time,
                fallback=booking.created_at,
            )
            earned += round(booking.price * percent / 100)
        complaint_state = complaint_status_for_percent(worker.default_percent, worker_penalties)
        complaint_effect = "Без снижения"
        if complaint_state.reduction_active and complaint_state.reduction_until is not None:
            complaint_effect = f"До {_format_datetime(complaint_state.reduction_until)}"
        payout = max(0, earned + worker.salary_base)
        total_payroll += payout
        payroll_rows.append([
            worker.name,
            "Да" if worker.active else "Нет",
            worker.default_percent,
            worker.salary_base,
            len(worker_completed),
            len(worker_active),
            earned,
            complaint_state.active_count,
            complaint_effect,
            payout,
        ])
    payroll_rows.sort(key=lambda item: (item[9], item[6], item[0]), reverse=True)
    client_rollup: dict[str, dict[str, Any]] = {}
    for booking in bookings:
        key = booking.client_id or f"{booking.client_name}|{booking.client_phone}"
        row = client_rollup.setdefault(
            key,
            {
                "name": booking.client_name,
                "phone": booking.client_phone,
                "vehicle": "",
                "total": 0,
                "completed": 0,
                "active": 0,
                "cancelled": 0,
                "revenue": 0,
                "first": None,
                "last": None,
            },
        )
        row["total"] += 1
        if booking.status == "completed":
            row["completed"] += 1
            row["revenue"] += booking.price
        elif booking.status in {"scheduled", "in_progress"}:
            row["active"] += 1
        elif booking.status == "cancelled":
            row["cancelled"] += 1
        vehicle = " / ".join(part for part in [booking.car or "", booking.plate or ""] if part)
        if vehicle and not row["vehicle"]:
            row["vehicle"] = vehicle
        dt = _booking_datetime(booking)
        if dt and (row["first"] is None or dt < row["first"]):
            row["first"] = dt
        if dt and (row["last"] is None or dt > row["last"]):
            row["last"] = dt
    client_rows = [
        [row["name"], row["phone"], row["vehicle"], row["total"], row["completed"], row["active"], row["cancelled"], row["revenue"], _format_datetime(row["first"]), _format_datetime(row["last"])]
        for row in client_rollup.values()
    ]
    client_rows.sort(key=lambda item: (item[7], item[3], item[0]), reverse=True)

    expense_rollup: dict[str, list[int]] = {}
    for expense in expenses:
        row = expense_rollup.setdefault(expense.category, [0, 0])
        row[0] += 1
        row[1] += expense.amount
    expense_category_rows = [
        [category, values[0], values[1], round((values[1] / total_expenses) * 100) if total_expenses else 0]
        for category, values in expense_rollup.items()
    ]
    expense_category_rows.sort(key=lambda item: (item[2], item[1], item[0]), reverse=True)
    expense_rows = [[expense.date, expense.category, expense.title, expense.amount, (expense.note or "").strip()] for expense in sorted(expenses, key=lambda item: (item.created_at, item.date), reverse=True)]

    stock_rows = [[item.category, item.name, item.qty, item.unit, item.unit_price, item.qty * item.unit_price, "Да" if item.qty <= 5 else "Нет"] for item in stock_items]
    stock_rows.sort(key=lambda item: (item[6] == "Да", item[5], item[1]), reverse=True)

    complaint_rows = []
    now = datetime.now().astimezone()
    for penalty in sorted(penalties, key=lambda item: item.created_at, reverse=True):
        active_until = penalty.active_until or penalty.created_at
        active_until_local = active_until.astimezone() if active_until.tzinfo is not None else active_until
        if penalty.revoked_at is not None:
            status = "Снята"
        elif (active_until_local < now) if active_until_local.tzinfo is not None else (active_until_local < now.replace(tzinfo=None)):
            status = "Истекла"
        else:
            status = "Активна"
        complaint_rows.append([
            penalty.worker.name if penalty.worker is not None else penalty.worker_id,
            penalty.title,
            penalty.reason,
            _format_datetime(penalty.created_at),
            _format_datetime(active_until),
            status,
            _format_datetime(penalty.revoked_at),
        ])

    booking_rows = [
        [
            _format_datetime(booking.created_at),
            _format_datetime(_booking_datetime(booking)),
            STATUS_LABELS.get(booking.status, booking.status),
            booking.client_name,
            booking.client_phone,
            " / ".join(part for part in [booking.car or "", booking.plate or ""] if part),
            booking.service,
            booking.box,
            PAYMENT_LABELS.get(booking.payment_type, booking.payment_type),
            booking.duration,
            booking.price,
            ", ".join(f"{link.worker_name} ({link.percent}%)" for link in booking.worker_links) or "Не назначены",
            (booking.notes or "").strip(),
        ]
        for booking in bookings
    ]

    metrics = [
        ExportMetric("Выручка", _format_money(revenue)),
        ExportMetric("Расходы", _format_money(total_expenses)),
        ExportMetric("Прибыль", _format_money(profit)),
        ExportMetric("Маржа", f"{margin}%"),
        ExportMetric("Средний чек", _format_money(avg_check)),
        ExportMetric("Всего записей", str(len(bookings))),
        ExportMetric("Завершено", str(len(completed))),
        ExportMetric("В работе", str(len(in_progress))),
        ExportMetric("Запланировано", str(len(scheduled))),
        ExportMetric("Отменено", str(len(cancelled))),
        ExportMetric("Конверсия в завершение", f"{completion_rate}%"),
        ExportMetric("Доля отмен", f"{cancellation_rate}%"),
        ExportMetric("К выплате сотрудникам", _format_money(total_payroll)),
        ExportMetric("Склад на сумму", _format_money(stock_value)),
        ExportMetric("Уникальных клиентов", str(len(client_rows))),
        ExportMetric("Активных жалоб", str(sum(1 for row in complaint_rows if row[5] == "Активна"))),
    ]

    return OwnerExportData(
        owner_name=owner.name.strip() or owner.login,
        company_name=company_name.strip() or "ATMOSFERA",
        generated_at=generated_at,
        period_from=period_from,
        period_to=period_to,
        metrics=metrics,
        service_rows=service_rows,
        payroll_rows=payroll_rows,
        client_rows=client_rows,
        booking_rows=booking_rows,
        expense_category_rows=expense_category_rows,
        expense_rows=expense_rows,
        stock_rows=stock_rows,
        complaint_rows=complaint_rows,
    )


def _render_excel_report(data: OwnerExportData) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Сводка"
    summary.merge_cells("A1:D1")
    summary["A1"] = data.company_name
    summary["A2"] = f"Владелец: {data.owner_name}"
    summary["A3"] = f"Сформирован: {data.generated_at.strftime('%d.%m.%Y %H:%M')}"
    summary["A4"] = f"Период записей: {data.period_from} — {data.period_to}"
    summary["A6"] = "Показатель"
    summary["B6"] = "Значение"
    for row_index, metric in enumerate(data.metrics, start=7):
        summary.cell(row=row_index, column=1, value=metric.label)
        summary.cell(row=row_index, column=2, value=metric.value)
    _style_heading(summary, "A1", "A2", "A3", "A4")
    _style_table(summary, 6, 7, 6 + len(data.metrics), 2)
    _autosize(summary)

    _append_sheet(workbook, "Услуги", ["Категория", "Услуга", "Всего", "Завершено", "Активно", "Отменено", "Выручка", "Средний чек", "Последняя запись"], data.service_rows, currency_cols={7, 8})
    _append_sheet(workbook, "Сотрудники", ["Сотрудник", "Активен", "%", "Оклад", "Завершено", "Активных задач", "Заработано", "Активные жалобы", "Эффект", "К выплате"], data.payroll_rows, currency_cols={4, 7, 10})
    _append_sheet(workbook, "Клиенты", ["Клиент", "Телефон", "Авто / номер", "Всего", "Завершено", "Активно", "Отменено", "Выручка", "Первая запись", "Последняя запись"], data.client_rows, currency_cols={8})
    _append_sheet(workbook, "Реестр записей", ["Создана", "Запись на", "Статус", "Клиент", "Телефон", "Авто / номер", "Услуга", "Бокс", "Оплата", "Длительность", "Стоимость", "Сотрудники", "Комментарий"], data.booking_rows, currency_cols={11})
    _append_sheet(workbook, "Расходы", ["Категория", "Операций", "Сумма", "Доля %"], data.expense_category_rows, currency_cols={3})
    _append_sheet(workbook, "Журнал расходов", ["Дата", "Категория", "Статья", "Сумма", "Примечание"], data.expense_rows, currency_cols={4})
    _append_sheet(workbook, "Склад", ["Категория", "Позиция", "Остаток", "Ед.", "Цена за ед.", "Стоимость", "Низкий остаток"], data.stock_rows, currency_cols={5, 6})
    _append_sheet(workbook, "Жалобы", ["Сотрудник", "Заголовок", "Причина", "Создана", "Активна до", "Статус", "Снята"], data.complaint_rows)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _render_owner_summary_excel_report(data: OwnerSummaryExportData) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Сводка"
    summary.merge_cells("A1:D1")
    summary["A1"] = data.company_name
    summary["A2"] = data.title
    summary["A3"] = f"Владелец: {data.owner_name}"
    summary["A4"] = f"Период: {data.period_label}"
    summary["A5"] = f"Направление: {data.segment_label}"
    summary["A6"] = f"Сформирован: {data.generated_at.strftime('%d.%m.%Y %H:%M')}"
    summary["A8"] = "Показатель"
    summary["B8"] = "Значение"
    for row_index, metric in enumerate(data.metrics, start=9):
        summary.cell(row=row_index, column=1, value=metric.label)
        summary.cell(row=row_index, column=2, value=metric.value)
    _style_heading(summary, "A1", "A2", "A3", "A4", "A5", "A6")
    _style_table(summary, 8, 9, 8 + len(data.metrics), 2)
    _autosize(summary)

    _append_sheet(
        workbook,
        "Статусы",
        ["Статус", "Количество", "Сумма записей", "Минут", "Доля %"],
        data.status_rows,
        currency_cols={3},
    )
    _append_sheet(
        workbook,
        "Оплаты",
        ["Способ оплаты", "Всего", "Завершено", "Выручка", "Открытая сумма"],
        data.payment_rows,
        currency_cols={4, 5},
    )
    _append_sheet(
        workbook,
        "По дням",
        ["Дата", "Всего", "Завершено", "Активно", "На уточнении", "Отменено", "Выручка", "Средний чек", "Минут", "Клиентов", "Боксов"],
        data.day_rows,
        currency_cols={7, 8},
    )
    _append_sheet(
        workbook,
        "По боксам",
        ["Бокс", "Всего", "Завершено", "Активно", "На уточнении", "Отменено", "Выручка", "Средний чек", "Минут", "Последняя запись"],
        data.box_rows,
        currency_cols={7, 8},
    )
    _append_sheet(
        workbook,
        "Услуги",
        ["Категория", "Услуга", "Всего", "Завершено", "Активно", "На уточнении", "Отменено", "Выручка", "Средний чек", "Последняя запись"],
        data.service_rows,
        currency_cols={8, 9},
    )
    _append_sheet(
        workbook,
        "Клиенты",
        ["Клиент", "Телефон", "Авто / номер", "Всего", "Завершено", "Активно", "На уточнении", "Отменено", "Выручка", "Первая запись", "Последняя запись"],
        data.client_rows,
        currency_cols={9},
    )
    _append_sheet(
        workbook,
        "Сотрудники",
        ["Сотрудник", "Всего", "Завершено", "Активно", "На уточнении", "Отменено", "Выручка", "Начислено", "Средний %", "Минут", "Последняя запись"],
        data.worker_rows,
        currency_cols={7, 8},
    )
    _append_sheet(
        workbook,
        "Реестр записей",
        ["Создана", "Запись на", "Статус", "Клиент", "Телефон", "Авто / номер", "Услуга", "Категория", "Бокс", "Оплата", "Длительность", "Стоимость", "Сотрудники", "Комментарий"],
        data.booking_rows,
        currency_cols={12},
    )
    _append_sheet(
        workbook,
        "Требуют внимания",
        ["Что сделать", "Статус", "Клиент", "Услуга", "Когда", "Бокс", "Сотрудники", "Комментарий"],
        data.action_rows,
    )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _append_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list[Any]], *, currency_cols: set[int] | None = None) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    if rows:
        for row in rows:
            sheet.append(row)
    else:
        sheet.append(["Нет данных", *([""] * (len(headers) - 1))])
    _style_table(sheet, 1, 2, sheet.max_row, len(headers))
    for row_index in range(2, sheet.max_row + 1):
        for column_index in currency_cols or set():
            _apply_currency(sheet.cell(row=row_index, column=column_index))
    _autosize(sheet)


def _render_pdf_report(data: OwnerExportData) -> bytes:
    buffer = io.BytesIO()
    font_name = _pdf_font_name()
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("OwnerTitle", parent=styles["Heading1"], fontName=font_name, fontSize=18, leading=22, textColor=colors.HexColor("#0B1226"))
    subtitle_style = ParagraphStyle("OwnerSubtitle", parent=styles["Normal"], fontName=font_name, fontSize=9, leading=11, textColor=colors.HexColor("#5B6470"))
    section_style = ParagraphStyle("OwnerSection", parent=styles["Heading2"], fontName=font_name, fontSize=12, leading=14, textColor=colors.HexColor("#0A84FF"))

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=f"Подробный отчёт владельца {data.company_name}",
        author=data.owner_name,
    )

    story: list[Any] = [
        Paragraph(_escape(data.company_name), title_style),
        Spacer(1, 3),
        Paragraph(_escape(f"Владелец: {data.owner_name} · Сформирован: {data.generated_at.strftime('%d.%m.%Y %H:%M')} · Период записей: {data.period_from} — {data.period_to}"), subtitle_style),
        Spacer(1, 6),
        _pdf_table([["Показатель", "Значение"], *[[item.label, item.value] for item in data.metrics]], font_name, "#0A84FF"),
        Spacer(1, 8),
    ]

    _pdf_section(story, section_style, font_name, "Услуги и выручка", ["Категория", "Услуга", "Всего", "Завершено", "Активно", "Отменено", "Выручка", "Средний чек", "Последняя запись"], data.service_rows)
    _pdf_section(story, section_style, font_name, "Сотрудники и выплаты", ["Сотрудник", "Активен", "%", "Оклад", "Завершено", "Активных задач", "Заработано", "Активные жалобы", "Эффект", "К выплате"], _format_rows(data.payroll_rows, currency_cols={4, 7, 10}))
    _pdf_section(story, section_style, font_name, "Клиенты", ["Клиент", "Телефон", "Авто / номер", "Всего", "Завершено", "Активно", "Отменено", "Выручка", "Первая запись", "Последняя запись"], _format_rows(data.client_rows, currency_cols={8}))
    _pdf_section(story, section_style, font_name, "Расходы по категориям", ["Категория", "Операций", "Сумма", "Доля %"], _format_rows(data.expense_category_rows, currency_cols={3}))
    _pdf_section(story, section_style, font_name, "Журнал расходов", ["Дата", "Категория", "Статья", "Сумма", "Примечание"], _format_rows(data.expense_rows, currency_cols={4}))
    _pdf_section(story, section_style, font_name, "Склад", ["Категория", "Позиция", "Остаток", "Ед.", "Цена за ед.", "Стоимость", "Низкий остаток"], _format_rows(data.stock_rows, currency_cols={5, 6}))
    _pdf_section(story, section_style, font_name, "Жалобы", ["Сотрудник", "Заголовок", "Причина", "Создана", "Активна до", "Статус", "Снята"], data.complaint_rows)
    _pdf_section(story, section_style, font_name, "Полный реестр записей", ["Создана", "Запись на", "Статус", "Клиент", "Телефон", "Авто / номер", "Услуга", "Бокс", "Оплата", "Длительность", "Стоимость", "Сотрудники", "Комментарий"], _format_rows(data.booking_rows, currency_cols={11}))

    doc.build(story)
    return buffer.getvalue()


def _pdf_section(story: list[Any], section_style: ParagraphStyle, font_name: str, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    story.append(Paragraph(title, section_style))
    story.append(Spacer(1, 3))
    table_rows = [headers, *([["Нет данных", *([""] * (len(headers) - 1))]] if not rows else rows)]
    story.append(_pdf_table(table_rows, font_name))
    story.append(Spacer(1, 8))


def _pdf_table(rows: list[list[Any]], font_name: str, header_color: str = "#0E1624") -> LongTable:
    normalized = [[Paragraph(_escape(str(cell)), _pdf_cell_style(font_name)) for cell in row] for row in rows]
    table = LongTable(normalized, repeatRows=1, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(header_color)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("LEADING", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F8FB")]),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D7DEE7")),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    return table


def _format_rows(rows: list[list[Any]], *, currency_cols: set[int]) -> list[list[Any]]:
    formatted: list[list[Any]] = []
    for row in rows:
        next_row = []
        for index, value in enumerate(row, start=1):
            if index in currency_cols and isinstance(value, (int, float)):
                next_row.append(_format_money(int(value)))
            elif value is None:
                next_row.append("")
            else:
                next_row.append(value)
        formatted.append(next_row)
    return formatted


def _style_heading(sheet, *cells: str) -> None:
    if cells:
        sheet[cells[0]].font = Font(size=16, bold=True, color="0B1226")
    for cell_name in cells[1:]:
        sheet[cell_name].font = Font(size=10, color="5B6470")


def _style_table(sheet, header_row: int, start_row: int, end_row: int, end_col: int) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="0A84FF")
    header_font = Font(bold=True, color="FFFFFF")
    zebra_fill = PatternFill(fill_type="solid", fgColor="F6F8FB")
    thin = Side(style="thin", color="D7DEE7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, end_col + 1):
        cell = sheet.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in range(start_row, end_row + 1):
        for col in range(1, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            if row % 2 == 0:
                cell.fill = zebra_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _apply_currency(cell) -> None:
    cell.number_format = '#,##0 "руб."'
    cell.alignment = Alignment(horizontal="right", vertical="center")


def _autosize(sheet) -> None:
    for column in sheet.columns:
        letter = get_column_letter(column[0].column)
        max_length = 0
        for cell in column:
            max_length = max(max_length, len("" if cell.value is None else str(cell.value)))
        sheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 36)


@lru_cache(maxsize=1)
def _pdf_font_name() -> str:
    candidates = [
        os.getenv("OWNER_EXPORT_FONT_PATH", ""),
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/calibri.ttf",
        "/usr/share/fonts/truetype/msttcorefonts/Arial.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/TTF/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    ]
    for raw_path in candidates:
        if not raw_path:
            continue
        candidate = Path(raw_path)
        if candidate.is_file():
            name = "OwnerExportFont"
            if name not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont(name, str(candidate)))
            return name
    return "Helvetica"


@lru_cache(maxsize=1)
def _pdf_cell_style(font_name: str) -> ParagraphStyle:
    return ParagraphStyle("OwnerExportCell", fontName=font_name, fontSize=7.5, leading=9, textColor=colors.HexColor("#111827"))


def _booking_datetime(booking: Booking) -> datetime | None:
    raw = f"{booking.date} {booking.time}".strip()
    for fmt in ("%d.%m.%Y %H:%M", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def _booking_sort_key(booking: Booking) -> tuple[datetime, datetime]:
    local_now = datetime.now().astimezone()
    booking_dt = _booking_datetime(booking)
    primary = _as_local_datetime(booking_dt, local_now) if booking_dt is not None else _as_local_datetime(booking.created_at, local_now)
    secondary = _as_local_datetime(booking.created_at, local_now)
    return primary, secondary


def _as_local_datetime(value: datetime, reference: datetime) -> datetime:
    target_tz = reference.tzinfo
    if value.tzinfo is None:
        return value.replace(tzinfo=target_tz)
    return value.astimezone(target_tz)


def _parse_date_for_sort(value: str) -> datetime:
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return datetime.max


def _format_datetime(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.astimezone().strftime("%d.%m.%Y %H:%M") if value.tzinfo is not None else value.strftime("%d.%m.%Y %H:%M")


def _format_money(value: int) -> str:
    return f"{value:,.0f}".replace(",", " ") + " руб."


def _escape(value: str) -> str:
    return escape(value).replace("\n", "<br/>")

